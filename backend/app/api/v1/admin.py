from __future__ import annotations
from datetime import datetime, timezone, date
import re
import copy
import csv
import io
import itertools
import json

from fastapi import APIRouter, Depends, HTTPException, Body, Query, UploadFile, File, Response, Request
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from sqlalchemy import select, func, desc, cast, Integer, update

from app.api.v1.deps import (
    get_db,
    get_current_user,
    get_allowed_organization_ids,
    get_allowed_location_ids,
    GLOBAL_ROLE_VALUES,
    user_has_any_role,
)
from app.models.submission import Submission
from app.services.rbac import require_roles, require_group_access, require_org_access
from app.models.role import Role
from app.models.organization import Organization
from app.models.location import Location
from app.models.survey import Survey, SurveyVersion
from app.models.user import User
from app.models.user_organization import UserOrganization
from app.models.token import UserRole
from app.models.stay import Stay
from app.api.v1.stays import router as stays_router
from app.models.group_survey_binding import GroupSurveyBinding
try:
    from app.models.audit_checklist import ChecklistTemplate, ChecklistQuestion  # type: ignore
except Exception:  # pragma: no cover
    ChecklistTemplate = None  # type: ignore
    ChecklistQuestion = None  # type: ignore
from app.services.public_url import build_public_url
from app.services.qr import make_qr_png, make_qr_svg
from app.services.humanize_answers import humanize_extra_answers
from app.services.review_links import (
    compute_effective_review_links,
    extract_location_review_links,
    extract_org_group_review_links,
    validate_review_links,
)

router = APIRouter(prefix="/admin", tags=["admin"])


router.include_router(stays_router)


def _clip_url(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # avoid storing huge payloads
    return s[:2048]


def _slugify(value: str) -> str:
    s = (value or "").strip().lower()
    s = s.replace("_", "-")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s


async def _make_unique_location_slug(db: AsyncSession, base: str) -> str:
    base = _slugify(base)
    if not base:
        raise HTTPException(status_code=400, detail="Cannot generate slug")

    slug = base
    for i in range(0, 50):
        exists = (await db.execute(select(Location.id).where(Location.slug == slug))).scalar_one_or_none()
        if not exists:
            return slug
        slug = f"{base}-{i+2}"

    raise HTTPException(status_code=409, detail="Cannot generate unique slug")


# Оставляем твой endpoint, чтобы ничего не упало
@router.get("/me")
async def me(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _user=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            Role.auditor,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.auditor_global,
        )
    ),
):
    roles = (
        await db.execute(select(UserRole).where(UserRole.user_id == user.id))
    ).scalars().all()

    is_global = any(r.role in GLOBAL_ROLE_VALUES for r in roles)

    allowed_org_ids = await get_allowed_organization_ids(db=db, user=user)
    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)

    locs = []
    if allowed_loc_ids:
        locs = (
            await db.execute(
                select(Location)
                .where(Location.id.in_(allowed_loc_ids), Location.is_active == True)  # noqa: E712
                .order_by(Location.organization_id, Location.name)
            )
        ).scalars().all()

    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "is_global": bool(is_global),
        "roles": [
            {"role": r.role, "organization_id": r.organization_id, "location_id": r.location_id}
            for r in roles
        ],
        "allowed_organization_ids": allowed_org_ids,
        "allowed_locations": [
            {
                "id": l.id,
                "organization_id": l.organization_id,
                "type": l.type,
                "code": l.code,
                "name": l.name,
                "slug": l.slug,
                "is_active": l.is_active,
            }
            for l in locs
        ],
    }


# =========================
# Organizations (Hotels)
# =========================

@router.get("/organizations")
async def list_organizations(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
        _=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            Role.auditor,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.auditor_global,
        )
    ),
):
    """
    Важно для админки CRUD:
      - для глобальных ролей возвращаем ВСЕ организации (включая is_active=False),
        чтобы можно было реактивировать
      - для scoped пользователей — только доступные активные
    """
    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)

    if is_global:
        orgs = (await db.execute(select(Organization).order_by(Organization.id))).scalars().all()
    else:
        allowed = await get_allowed_organization_ids(db=db, user=user)
        if not allowed:
            return []
        orgs = (
            await db.execute(
                select(Organization)
                .where(Organization.id.in_(allowed))
                .order_by(Organization.id)
            )
        ).scalars().all()

    return [{"id": o.id, "name": o.name, "slug": o.slug, "is_active": o.is_active} for o in orgs]


@router.get("/organizations/{organization_id}")
async def get_organization(
    organization_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
        Role.auditor_global,
        Role.auditor,
    )),
):
    """
    Single organization fetch:
      - global roles: can read any org (including inactive)
      - non-global: only allowed orgs AND only if org.is_active
    """
    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)

    if not is_global:
        allowed_orgs = await get_allowed_organization_ids(db=db, user=user)
        if organization_id not in allowed_orgs:
            raise HTTPException(status_code=403, detail="No access to this organization")
        if not org.is_active:
            # non-global users don't see deactivated orgs
            raise HTTPException(status_code=404, detail="Organization not found")

    return {"id": org.id, "name": org.name, "slug": org.slug, "is_active": org.is_active}


@router.get("/locations/{location_id}")
async def get_location(
    location_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
        Role.auditor_global,
        Role.auditor,
    )),
):
    """
    Single location fetch:
      - global roles: can read any location (including inactive)
      - non-global:
          * must have access to its organization
          * if org-wide manager/auditor -> any ACTIVE location in org
          * else -> only if location_id in allowed_location_ids
          * inactive locations hidden from non-global
    """
    loc = (await db.execute(select(Location).where(Location.id == location_id))).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)

    if not is_global:
        # hide inactive from non-global
        if not loc.is_active:
            raise HTTPException(status_code=404, detail="Location not found")

        allowed_orgs = await get_allowed_organization_ids(db=db, user=user)
        if loc.organization_id not in allowed_orgs:
            raise HTTPException(status_code=403, detail="No access to this location")

        # org-wide roles check (manager/auditor with no location_id)
        role_rows = (
            await db.execute(
                select(UserRole.role, UserRole.organization_id, UserRole.location_id)
                .where(UserRole.user_id == user.id)
            )
        ).all()

        has_org_wide = any(
            (r[0] in {Role.manager.value, Role.auditor.value})
            and (r[1] is None or r[1] == loc.organization_id)
            and (r[2] is None)
            for r in role_rows
        )

        if not has_org_wide:
            allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
            if loc.id not in allowed_loc_ids:
                raise HTTPException(status_code=403, detail="No access to this location")

    return {
        "id": loc.id,
        "organization_id": loc.organization_id,
        "type": loc.type,
        "code": loc.code,
        "name": loc.name,
        "slug": loc.slug,
        "is_active": loc.is_active,
    }


async def _get_location_with_access_check(
    *,
    db: AsyncSession,
    user: User,
    location_id: int,
) -> Location:
    """Fetch location and apply the same access rules as get_location()."""
    loc = (await db.execute(select(Location).where(Location.id == location_id))).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)
    if not is_global:
        if not loc.is_active:
            raise HTTPException(status_code=404, detail="Location not found")

        allowed_orgs = await get_allowed_organization_ids(db=db, user=user)
        if loc.organization_id not in allowed_orgs:
            raise HTTPException(status_code=403, detail="No access to this location")

        # org-wide roles check (manager/auditor with no location_id)
        role_rows = (
            await db.execute(
                select(UserRole.role, UserRole.organization_id, UserRole.location_id)
                .where(UserRole.user_id == user.id)
            )
        ).all()

        has_org_wide = any(
            (r[0] in {Role.manager.value, Role.auditor.value})
            and (r[1] is None or r[1] == loc.organization_id)
            and (r[2] is None)
            for r in role_rows
        )

        if not has_org_wide:
            allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
            if loc.id not in allowed_loc_ids:
                raise HTTPException(status_code=403, detail="No access to this location")

    return loc


# =======================
# Review links (Yandex / 2GIS) settings
# Storage:
#   Organization.settings["review_links_by_group"][group_key] = { yandex_url, twogis_url }
#   Location.settings["review_links"] = { inherit: bool, yandex_url, twogis_url }
# Inheritance:
#   location override (inherit=False) > org group default > org default > none
# =======================

def _normalize_group_key(group_key: str) -> str:
    g = (group_key or "").strip()
    if not g or len(g) > 32:
        raise HTTPException(status_code=400, detail="Invalid group_key")
    return g


async def _service_manager_needs_group_check(db: AsyncSession, user: User) -> bool:
    """True only when user is 'service_manager-only' (no higher roles)."""
    has_sm = await user_has_any_role(db, user, [Role.service_manager.value])
    if not has_sm:
        return False

    higher_roles = [
        Role.admin.value,
        Role.ops_director.value,
        Role.manager.value,
        Role.director.value,
        Role.super_admin.value,
        Role.auditor.value,
        Role.auditor_global.value,
    ]
    has_higher = await user_has_any_role(db, user, higher_roles)
    return not has_higher


@router.get("/organizations/{organization_id}/groups/{group_key}/review-links")
async def get_group_review_links(
    organization_id: int,
    group_key: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            Role.auditor,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.auditor_global,
            Role.super_admin,
        )
    ),
):
    group_key = _normalize_group_key(group_key)

    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    # Base org access for everyone
    await require_org_access(db=db, user=user, organization_id=organization_id)

    # service_manager должен иметь реальный доступ к группе
    if await _service_manager_needs_group_check(db, user):
        allowed_loc_ids = await require_group_access(
            db=db, user=user, organization_id=organization_id, group_key=group_key
        )
        if not allowed_loc_ids:
            raise HTTPException(status_code=403, detail="No access to this group")

    links = extract_org_group_review_links(getattr(org, "settings", None), group_key)
    errors = validate_review_links(links)

    return {
        "organization_id": organization_id,
        "group_key": group_key,
        "yandex_url": links.get("yandex_url"),
        "twogis_url": links.get("twogis_url"),
        "errors": errors,
    }


@router.patch("/organizations/{organization_id}/groups/{group_key}/review-links")
async def patch_group_review_links(
    organization_id: int,
    group_key: str,
    payload: dict = Body(default={}),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.super_admin,
        )
    ),
):
    group_key = _normalize_group_key(group_key)

    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    await require_org_access(db=db, user=user, organization_id=organization_id)

    if await _service_manager_needs_group_check(db, user):
        allowed_loc_ids = await require_group_access(
            db=db, user=user, organization_id=organization_id, group_key=group_key
        )
        if not allowed_loc_ids:
            raise HTTPException(status_code=403, detail="No access to this group")

    settings = dict(getattr(org, "settings", None) or {})
    by_group = dict(settings.get("review_links_by_group") or {})
    existing = dict(by_group.get(group_key) or {})

    if "yandex_url" in payload:
        existing["yandex_url"] = _clip_url(payload.get("yandex_url"))
    if "twogis_url" in payload:
        existing["twogis_url"] = _clip_url(payload.get("twogis_url"))

    by_group[group_key] = existing
    settings["review_links_by_group"] = by_group

    org.settings = settings  # <-- важно
    await db.commit()
    await db.refresh(org)

    y = existing.get("yandex_url")
    t = existing.get("twogis_url")

    if "yandex_url" in payload:
        y = _clip_url(payload.get("yandex_url"))
    if "twogis_url" in payload:
        t = _clip_url(payload.get("twogis_url"))

    by_group[group_key] = {"yandex_url": y, "twogis_url": t}
    settings["review_links_by_group"] = by_group
    org.settings = settings

    await db.commit()
    await db.refresh(org)

    links = extract_org_group_review_links(getattr(org, "settings", None), group_key)
    errors = validate_review_links(links)

    return {
        "organization_id": organization_id,
        "group_key": group_key,
        "yandex_url": links.get("yandex_url"),
        "twogis_url": links.get("twogis_url"),
        "errors": errors,
    }


@router.get("/locations/{location_id}/review-links")
async def get_location_review_links(
    location_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            Role.auditor,
            Role.manager,
            Role.director,
            Role.auditor_global,
            Role.super_admin,
        )
    ),
):
    loc = await _get_location_with_access_check(db=db, user=user, location_id=location_id)

    org = (await db.execute(select(Organization).where(Organization.id == loc.organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    loc_cfg = extract_location_review_links(getattr(loc, "settings", None))
    grp = extract_org_group_review_links(getattr(org, "settings", None), loc.type)

    effective = compute_effective_review_links(
        org_settings=getattr(org, "settings", None),
        group_key=loc.type,
        location_settings=getattr(loc, "settings", None),
    )

    override_errors = validate_review_links(
        {"yandex_url": loc_cfg.get("yandex_url"), "twogis_url": loc_cfg.get("twogis_url")}
    )

    return {
        "location_id": loc.id,
        "organization_id": loc.organization_id,
        "group_key": loc.type,
        "inherit": bool(loc_cfg.get("inherit")),
        "override": {
            "yandex_url": loc_cfg.get("yandex_url"),
            "twogis_url": loc_cfg.get("twogis_url"),
        },
        "group_default": {
            "yandex_url": grp.get("yandex_url"),
            "twogis_url": grp.get("twogis_url"),
        },
        "effective": effective,
        "errors": override_errors,
    }


@router.patch("/locations/{location_id}/review-links")
async def patch_location_review_links(
    location_id: int,
    payload: dict = Body(default={}),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            Role.manager,
            Role.director,
            Role.super_admin,
        )
    ),
):
    loc = await _get_location_with_access_check(db=db, user=user, location_id=location_id)

    org = (await db.execute(select(Organization).where(Organization.id == loc.organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    settings = dict(getattr(loc, "settings", None) or {})
    review_links = dict(settings.get("review_links") or {})

    if "inherit" in payload:
        review_links["inherit"] = bool(payload.get("inherit"))
    if "yandex_url" in payload:
        review_links["yandex_url"] = _clip_url(payload.get("yandex_url"))
    if "twogis_url" in payload:
        review_links["twogis_url"] = _clip_url(payload.get("twogis_url"))

    settings["review_links"] = review_links
    loc.settings = settings  # <-- важно
    await db.commit()
    await db.refresh(loc)

    loc_cfg = extract_location_review_links(getattr(loc, "settings", None))
    grp = extract_org_group_review_links(getattr(org, "settings", None), loc.type)

    effective = compute_effective_review_links(
        org_settings=getattr(org, "settings", None),
        group_key=loc.type,
        location_settings=getattr(loc, "settings", None),
    )

    override_errors = validate_review_links(
        {"yandex_url": loc_cfg.get("yandex_url"), "twogis_url": loc_cfg.get("twogis_url")}
    )

    return {
        "location_id": loc.id,
        "organization_id": loc.organization_id,
        "group_key": loc.type,
        "inherit": bool(loc_cfg.get("inherit")),
        "override": {
            "yandex_url": loc_cfg.get("yandex_url"),
            "twogis_url": loc_cfg.get("twogis_url"),
        },
        "group_default": {
            "yandex_url": grp.get("yandex_url"),
            "twogis_url": grp.get("twogis_url"),
        },
        "effective": effective,
        "errors": override_errors,
    }


@router.get("/locations/{location_id}/public-url")
async def get_location_public_url(
    location_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            Role.admin,
            Role.ops_director,
            Role.manager,
            Role.service_manager,
            Role.director,
            Role.auditor_global,
            Role.auditor,
        )
    ),
):
    """Return stable public URL for a location (based on its slug)."""
    loc = await _get_location_with_access_check(db=db, user=user, location_id=location_id)
    return {"public_url": build_public_url(loc.slug)}


def _cache_headers(etag: str, *, filename: str) -> dict[str, str]:
    # 1 day caching; revalidate with ETag.
    return {
        "ETag": f"\"{etag}\"",
        "Cache-Control": "public, max-age=86400",
        "Content-Disposition": f"attachment; filename=\"{filename}\"",
    }


@router.get("/locations/{location_id}/qr.svg")
async def get_location_qr_svg(
    location_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            Role.admin,
            Role.ops_director,
            Role.manager,
            Role.service_manager,
            Role.director,
            Role.auditor_global,
            Role.auditor,
        )
    ),
):
    """Download location QR as SVG (static as long as slug + PUBLIC_BASE_URL are stable)."""
    loc = await _get_location_with_access_check(db=db, user=user, location_id=location_id)
    payload = build_public_url(loc.slug)
    svg_bytes, etag = make_qr_svg(payload)

    inm = (request.headers.get("if-none-match") or "").strip().strip('"')
    if inm and inm == etag:
        return Response(status_code=304, headers=_cache_headers(etag, filename=f"location-{loc.id}-{loc.slug}.svg"))

    return Response(
        content=svg_bytes,
        media_type="image/svg+xml",
        headers=_cache_headers(etag, filename=f"location-{loc.id}-{loc.slug}.svg"),
    )


@router.get("/locations/{location_id}/qr.png")
async def get_location_qr_png(
    location_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            Role.admin,
            Role.ops_director,
            Role.manager,
            Role.service_manager,
            Role.director,
            Role.auditor_global,
            Role.auditor,
        )
    ),
):
    """Download location QR as PNG (optional; useful for screens/devices)."""
    loc = await _get_location_with_access_check(db=db, user=user, location_id=location_id)
    payload = build_public_url(loc.slug)

    try:
        png_bytes, etag = make_qr_png(payload)
    except Exception:
        raise HTTPException(status_code=501, detail="PNG QR is not available in this build")

    inm = (request.headers.get("if-none-match") or "").strip().strip('"')
    if inm and inm == etag:
        return Response(status_code=304, headers=_cache_headers(etag, filename=f"location-{loc.id}-{loc.slug}.png"))

    return Response(
        content=png_bytes,
        media_type="image/png",
        headers=_cache_headers(etag, filename=f"location-{loc.id}-{loc.slug}.png"),
    )


@router.post("/organizations")
async def create_organization(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles(Role.director)),  # создавать отели может директор (глобальный)
):
    """
    Создать новую организацию (отель).
    payload: { "name": "Hotel X", "slug": "hotel-x" }
    """
    name = (payload.get("name") or "").strip()
    slug = (payload.get("slug") or "").strip()

    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    if not slug:
        raise HTTPException(status_code=400, detail="slug is required")

    org = Organization(name=name, slug=slug, is_active=True)
    db.add(org)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Organization slug already exists")

    await db.refresh(org)
    return {"id": org.id, "name": org.name, "slug": org.slug, "is_active": org.is_active}


@router.patch("/organizations/{organization_id}")
async def update_organization(
    organization_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles(Role.director)),
):
    """
    payload (any subset):
      { "name": "...", "slug": "...", "is_active": true|false }
    """
    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    if "name" in payload:
        name = (payload.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="name cannot be empty")
        org.name = name

    if "slug" in payload:
        slug = (payload.get("slug") or "").strip()
        if not slug:
            raise HTTPException(status_code=400, detail="slug cannot be empty")
        org.slug = _slugify(slug) or slug

    if "is_active" in payload:
        org.is_active = bool(payload.get("is_active"))

    db.add(org)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Organization slug already exists")

    await db.refresh(org)
    return {"id": org.id, "name": org.name, "slug": org.slug, "is_active": org.is_active}



@router.post("/organizations/{organization_id}/grant")
async def grant_user_to_organization(
    organization_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles(Role.director)),  # доступы выдаёт директор
):
    """
    Выдать пользователю доступ к организации (отелю).
    payload: { "user_id": 123 }
    """
    try:
        user_id = int(payload["user_id"])
    except Exception:
        raise HTTPException(status_code=400, detail="user_id is required")

    org = (
        await db.execute(select(Organization).where(Organization.id == organization_id))
    ).scalar_one_or_none()
    if not org or not org.is_active:
        raise HTTPException(status_code=404, detail="Organization not found")

    user = (
        await db.execute(select(User).where(User.id == user_id))
    ).scalar_one_or_none()
    if not user or not user.is_active:
        raise HTTPException(status_code=404, detail="User not found")

    link = (
        await db.execute(
            select(UserOrganization).where(
                UserOrganization.user_id == user_id,
                UserOrganization.organization_id == organization_id,
            )
        )
    ).scalar_one_or_none()

    if link:
        link.is_active = True
        db.add(link)
        await db.commit()
        return {"ok": True, "user_id": user_id, "organization_id": organization_id, "status": "reactivated"}

    link = UserOrganization(user_id=user_id, organization_id=organization_id, is_active=True)
    db.add(link)
    await db.commit()
    return {"ok": True, "user_id": user_id, "organization_id": organization_id, "status": "granted"}


# =========================
# Locations (Rooms/Restaurants/Halls)
# =========================

@router.get("/organizations/{organization_id}/locations")
async def list_locations(
    organization_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
        Role.auditor_global,
        Role.auditor,
    )),
):
    """
    RBAC:
      - director/auditor_global -> все локации org (включая inactive)
      - manager/auditor (org-wide) -> все ACTIVE локации org (если org доступна)
      - service_manager/employee -> только назначенные ACTIVE локации в рамках org
    """
    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)

    if is_global:
        q = select(Location).where(Location.organization_id == organization_id)
    else:
        allowed_orgs = await get_allowed_organization_ids(db=db, user=user)
        if organization_id not in allowed_orgs:
            raise HTTPException(status_code=403, detail="No access to this organization")

        role_rows = (
            await db.execute(
                select(UserRole.role, UserRole.organization_id, UserRole.location_id)
                .where(UserRole.user_id == user.id)
            )
        ).all()

        has_org_wide = any(
            (r[0] in {Role.manager.value, Role.auditor.value})
            and (r[1] is None or r[1] == organization_id)
            and (r[2] is None)
            for r in role_rows
        )

        if has_org_wide:
            q = select(Location).where(
                Location.organization_id == organization_id,
                Location.is_active == True,  # noqa: E712
            )
        else:
            allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
            if not allowed_loc_ids:
                return []
            q = select(Location).where(
                Location.organization_id == organization_id,
                Location.id.in_(allowed_loc_ids),
                Location.is_active == True,  # noqa: E712
            )

    locs = (await db.execute(q.order_by(Location.name))).scalars().all()

    return [
        {
            "id": l.id,
            "organization_id": l.organization_id,
            "type": l.type,
            "code": l.code,
            "name": l.name,
            "slug": l.slug,
            "is_active": l.is_active,
        }
        for l in locs
    ]


@router.post("/locations")
async def create_location(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            Role.admin,
            Role.ops_director,
            # legacy compat
            Role.director,
            Role.super_admin,
            detail="No permission to create locations",
        )
    ),
):
    """
    payload:
    {
      "organization_id": 1,
      "type": "room|restaurant|conference_hall|banquet_hall|other",
      "code": "301",
      "name": "Номер 301",
      "slug": "demo-hotel-301"   # optional; если не задан — генерим {org_slug}-{code}
    }
    """
    try:
        organization_id = int(payload["organization_id"])
    except Exception:
        raise HTTPException(status_code=400, detail="organization_id is required")

    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    # Enforce org-scope for non-global roles (and keep behavior explicit)
    await require_org_access(db=db, user=user, organization_id=organization_id)

    loc_type = (payload.get("type") or "room").strip()
    code = (payload.get("code") or "").strip()
    name = (payload.get("name") or "").strip()
    slug = (payload.get("slug") or "").strip()

    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    if slug:
        slug = _slugify(slug)
        if not slug:
            raise HTTPException(status_code=400, detail="Invalid slug")
    else:
        suffix_src = code or name
        suffix = _slugify(suffix_src)
        if not suffix:
            raise HTTPException(status_code=400, detail="code or name must be non-empty to generate slug")
        base = f"{org.slug}-{suffix}"
        slug = await _make_unique_location_slug(db, base)

    loc = Location(
        organization_id=organization_id,
        type=loc_type,
        code=code,
        name=name,
        slug=slug,
        is_active=True,
    )

    db.add(loc)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Location slug already exists")

    await db.refresh(loc)
    return {
        "id": loc.id,
        "organization_id": loc.organization_id,
        "type": loc.type,
        "code": loc.code,
        "name": loc.name,
        "slug": loc.slug,
        "is_active": loc.is_active,
    }


# =========================
# Locations Import (Excel)
# =========================

_ALLOWED_LOCATION_TYPES: dict[str, str] = {
    # EN (canonical)
    "room": "room",
    "rooms": "room",
    "restaurant": "restaurant",
    "restaurants": "restaurant",
    "conference_hall": "conference_hall",
    "conference-hall": "conference_hall",
    "conference": "conference_hall",
    "hall": "conference_hall",
    "banquet_hall": "banquet_hall",
    "banquet-hall": "banquet_hall",
    "banquet": "banquet_hall",
    "other": "other",

    # RU
    "номер": "room",
    "номера": "room",
    "комната": "room",
    "комнаты": "room",

    "ресторан": "restaurant",
    "рестораны": "restaurant",

    "конференцзал": "conference_hall",
    "конференц_зал": "conference_hall",
    "конференц-зал": "conference_hall",
    "конференция": "conference_hall",
    "зал": "conference_hall",

    "банкет": "banquet_hall",
    "банкетныйзал": "banquet_hall",
    "банкетный_зал": "banquet_hall",
    "банкетный-зал": "banquet_hall",

    "другое": "other",
}


def _norm_loc_type(raw: str) -> str:
    key = (raw or "").strip().lower()
    key = key.replace("-", "_")
    key = key.replace(" ", "_")
    key = key.replace("__", "_")
    return _ALLOWED_LOCATION_TYPES.get(key, "")


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _extract_room_number(name: str) -> str:
    m = re.search(r"\d+", name or "")
    return m.group(0) if m else ""


def _generate_code(loc_type: str, name: str, used_codes: set[str]) -> str:
    """
    Deterministic-ish code generation.
    - room: first number in name (e.g. "Room 101" -> "101") else "ROOM"
    - others: <TYPE>-<SLUGBASE>
    Ensures uniqueness inside organization for *new* records by suffixing -2, -3...
    """
    if loc_type == "room":
        base = _extract_room_number(name) or "ROOM"
        cand = base
    else:
        slug_base = _slugify(name)
        if not slug_base:
            import hashlib
            slug_base = hashlib.md5((name or "").strip().lower().encode("utf-8")).hexdigest()[:10]

        prefix = (loc_type or "loc")[:3].upper()
        cand = f"{prefix}-{slug_base.upper()[:24]}"

    code = cand
    if code and code in used_codes:
        for i in range(2, 200):
            code2 = f"{cand}-{i}"
            if code2 not in used_codes:
                code = code2
                break
    if code:
        used_codes.add(code)
    return code


async def _ensure_import_access(db: AsyncSession, user: User, organization_id: int) -> Organization:
    org = (
        await db.execute(select(Organization).where(Organization.id == organization_id))
    ).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    allowed = await get_allowed_organization_ids(db=db, user=user)
    if organization_id not in allowed:
        raise HTTPException(status_code=403, detail="No access to this organization")

    return org


@router.get("/organizations/{organization_id}/locations-import/template")
async def download_locations_import_template(
    organization_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.director, Role.super_admin)),
):
    await _ensure_import_access(db=db, user=user, organization_id=organization_id)

    try:
        from openpyxl import Workbook  # lazy import
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Локации"
    ws.append(["Тип", "Название", "Код", "Слаг"])
    ws.append(["Номер", "Номер 101", "", ""])
    ws.append(["Ресторан", "Ресторан A", "", ""])
    ws.append(["Конференц-зал", "Зал 1", "", ""])

    buf = io.BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="locations_import_template.xlsx"'},
    )


@router.post("/organizations/{organization_id}/locations-import")
async def import_locations_from_excel(
    organization_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.director, Role.super_admin)),
):
    org = await _ensure_import_access(db=db, user=user, organization_id=organization_id)

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        from openpyxl import load_workbook  # lazy import
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    try:
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid xlsx file")

    if "locations" in wb.sheetnames:
        ws = wb["locations"]
    elif "Локации" in wb.sheetnames:
        ws = wb["Локации"]
    else:
        ws = wb.active

    # find first non-empty row as header
    header_row_idx = None
    header_values = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(_cell_str(c) for c in row):
            header_row_idx = i
            header_values = [(_cell_str(c) or "").strip() for c in row]
            break

    if header_row_idx is None or header_values is None:
        raise HTTPException(status_code=400, detail="No header row found")

    _HEADER_ALIASES: dict[str, str] = {
        # canonical
        "type": "type",
        "name": "name",
        "code": "code",
        "slug": "slug",

        # RU headers
        "тип": "type",
        "название": "name",
        "наименование": "name",
        "код": "code",
        "слаг": "slug",
        "slug": "slug",
    }

    header_norm_raw = [
        (_cell_str(h) or "")
        .strip()
        .lower()
        .replace("-", "_")
        .replace(" ", "_")
        for h in header_values
    ]

    header_keys = [_HEADER_ALIASES.get(h, h) for h in header_norm_raw]

    def col_idx(name: str) -> int | None:
        try:
            return header_keys.index(name)
        except ValueError:
            return None

    i_type = col_idx("type")
    i_name = col_idx("name")
    i_code = col_idx("code")
    i_slug = col_idx("slug")

    if i_type is None or i_name is None:
        raise HTTPException(status_code=400, detail="Required columns: Тип/Название (или type/name)")

    existing_locs = (
        await db.execute(select(Location).where(Location.organization_id == organization_id))
    ).scalars().all()

    by_slug: dict[str, Location] = {l.slug: l for l in existing_locs}
    by_code: dict[str, list[Location]] = {}
    used_codes: set[str] = set()
    for l in existing_locs:
        if l.code:
            used_codes.add(l.code)
            by_code.setdefault(l.code, []).append(l)

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []

    seen_pairs: set[tuple[str, str]] = set()

    for excel_row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row or not any(_cell_str(c) for c in row):
            continue

        raw_type = _cell_str(row[i_type] if i_type < len(row) else "")
        raw_name = _cell_str(row[i_name] if i_name < len(row) else "")
        raw_code = _cell_str(row[i_code] if (i_code is not None and i_code < len(row)) else "")
        raw_slug = _cell_str(row[i_slug] if (i_slug is not None and i_slug < len(row)) else "")

        if not raw_type or not raw_name:
            errors.append({"row": excel_row_idx, "error": "type and name are required"})
            continue

        loc_type = _norm_loc_type(raw_type)
        if not loc_type:
            errors.append({"row": excel_row_idx, "error": f"invalid type: {raw_type}"})
            continue

        name = raw_name.strip()
        pair_key = (loc_type, name.lower())
        if pair_key in seen_pairs:
            skipped += 1
            continue
        seen_pairs.add(pair_key)

        slug_in = _slugify(raw_slug) if raw_slug else ""
        code_in = raw_code.strip()

        loc: Location | None = None

        if slug_in:
            loc = by_slug.get(slug_in)
            if not loc:
                # protect public: slug must be globally unique
                other_id = (
                    await db.execute(select(Location.id).where(Location.slug == slug_in))
                ).scalar_one_or_none()
                if other_id:
                    errors.append({"row": excel_row_idx, "error": f"slug already exists: {slug_in}"})
                    continue

        elif code_in:
            cands = by_code.get(code_in, [])
            if len(cands) == 1:
                loc = cands[0]
            elif len(cands) > 1:
                errors.append({"row": excel_row_idx, "error": f"ambiguous code (multiple locations): {code_in}"})
                continue

        # generate code for new records
        if not code_in and not loc:
            code_in = _generate_code(loc_type=loc_type, name=name, used_codes=used_codes)

        # generate slug if missing
        if not slug_in:
            suffix_src = code_in or name
            suffix = _slugify(suffix_src)
            if not suffix:
                errors.append({"row": excel_row_idx, "error": "Cannot generate slug (empty code/name)"})
                continue
            base = f"{org.slug}-{suffix}"
            slug_in = await _make_unique_location_slug(db, base)

        # stable re-import: match by generated slug
        if not loc:
            loc = by_slug.get(slug_in)

        if loc:
            changed = False
            if loc.name != name:
                loc.name = name
                changed = True
            if loc.type != loc_type:
                loc.type = loc_type
                changed = True
            if code_in and loc.code != code_in:
                loc.code = code_in
                changed = True
            if not loc.is_active:
                loc.is_active = True
                changed = True

            if changed:
                db.add(loc)
                updated += 1
            else:
                skipped += 1
            continue

        new_loc = Location(
            organization_id=organization_id,
            type=loc_type,
            code=code_in or "",
            name=name,
            slug=slug_in,
            is_active=True,
        )
        db.add(new_loc)
        created += 1

    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Import failed due to unique constraint конфликт (slug)")

    return {
        "ok": True,
        "organization_id": organization_id,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


# =========================
# Audit checklist templates import (Excel)
# =========================

def _audit_bool(v, default: bool = False) -> bool:
    s = (_cell_str(v) or "").strip().lower()
    if s in {"1", "true", "yes", "y", "on", "да"}:
        return True
    if s in {"0", "false", "no", "n", "off", "нет"}:
        return False
    return default


def _audit_int(v, default: int = 0) -> int:
    s = (_cell_str(v) or "").strip()
    if not s:
        return default
    try:
        return int(float(s))
    except Exception:
        return default


def _audit_parse_meta(wb) -> dict:
    meta: dict[str, str] = {}
    for name in ("meta", "Meta", "META", "Мета", "мета"):
        if name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 2:
                    continue
                k = (_cell_str(row[0]) or "").strip().lower()
                v = (_cell_str(row[1]) or "").strip()
                if k:
                    meta[k] = v
            break
    return meta


def _audit_find_questions_sheet(wb):
    for name in ("questions", "Questions", "Вопросы", "вопросы"):
        if name in wb.sheetnames:
            return wb[name]
    return None


def _audit_parse_structured_questions(ws) -> tuple[str | None, list[dict]]:
    """
    Structured sheet expected columns (any aliases):
      - section / раздел / зона
      - question / вопрос
      - answer_type / тип
      - required / обязателен
      - allow_comment / комментарий
      - allow_photos / фото
      - order / порядок
      - options_json / опции
    """
    header_row_idx = None
    header_values = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(_cell_str(c) for c in row):
            header_row_idx = i
            header_values = [(_cell_str(c) or "").strip() for c in row]
            break

    if header_row_idx is None or header_values is None:
        raise HTTPException(status_code=400, detail="No header row found in questions sheet")

    aliases = {
        "section": "section",
        "раздел": "section",
        "зона": "section",

        "question": "question",
        "вопрос": "question",
        "пункт": "question",

        "answer_type": "answer_type",
        "type": "answer_type",
        "тип": "answer_type",

        "required": "required",
        "обязателен": "required",
        "обязательно": "required",

        "allow_comment": "allow_comment",
        "комментарий": "allow_comment",
        "comments": "allow_comment",

        "allow_photos": "allow_photos",
        "фото": "allow_photos",
        "photo": "allow_photos",
        "photos": "allow_photos",

        "order": "order",
        "порядок": "order",
        "№": "order",
        "#": "order",

        "options_json": "options_json",
        "options": "options_json",
        "опции": "options_json",
        "варианты": "options_json",
    }

    norm = []
    for h in header_values:
        key = (h or "").strip().lower()
        key = key.replace("-", "_").replace(" ", "_")
        norm.append(aliases.get(key, key))

    def col(name: str) -> int | None:
        try:
            return norm.index(name)
        except ValueError:
            return None

    i_section = col("section")
    i_question = col("question")
    i_answer_type = col("answer_type")
    i_required = col("required")
    i_allow_comment = col("allow_comment")
    i_allow_photos = col("allow_photos")
    i_order = col("order")
    i_options = col("options_json")

    if i_question is None:
        raise HTTPException(status_code=400, detail="questions sheet: missing required column 'question/вопрос'")

    rows_out: list[dict] = []
    for excel_row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row or not any(_cell_str(c) for c in row):
            continue

        q_text = _cell_str(row[i_question] if i_question < len(row) else "")
        if not q_text:
            continue

        section = _cell_str(row[i_section] if (i_section is not None and i_section < len(row)) else "")
        answer_type = _cell_str(row[i_answer_type] if (i_answer_type is not None and i_answer_type < len(row)) else "") or "yesno_score"

        is_required = _audit_bool(row[i_required] if (i_required is not None and i_required < len(row)) else "", default=False)
        allow_comment = _audit_bool(row[i_allow_comment] if (i_allow_comment is not None and i_allow_comment < len(row)) else "", default=True)
        allow_photos = _audit_bool(row[i_allow_photos] if (i_allow_photos is not None and i_allow_photos < len(row)) else "", default=True)
        order = _audit_int(row[i_order] if (i_order is not None and i_order < len(row)) else "", default=0)

        options = None
        if i_options is not None and i_options < len(row):
            raw_opt = _cell_str(row[i_options])
            if raw_opt:
                try:
                    options = json.loads(raw_opt)
                except Exception:
                    raise HTTPException(status_code=400, detail=f"Invalid options_json at row {excel_row_idx}")

        rows_out.append({
            "section": section,
            "text": q_text,
            "answer_type": answer_type,
            "options": options,
            "is_required": bool(is_required),
            "allow_comment": bool(allow_comment),
            "allow_photos": bool(allow_photos),
            "order": int(order),
        })

    return None, rows_out


def _audit_parse_hsk_like(ws) -> tuple[str | None, list[dict]]:
    """
    Supports Word-like table exported to xlsx:
      columns like:
        ['', 'Зона', 'Балл ...', 'Комментарий']
      or:
        ['Зона', 'Балл ...', 'Комментарий']
    Section rows typically have only section cell filled, questions are in 'Зона' col.
    """
    header_row_idx = None
    header = None

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [(_cell_str(c) or "") for c in (row or [])]
        if not any(cells):
            continue
        joined = " | ".join([c.lower() for c in cells if c])
        if ("зона" in joined) and ("балл" in joined):
            header_row_idx = i
            header = cells
            break

    if header_row_idx is None or header is None:
        raise HTTPException(status_code=400, detail="Cannot detect header row (expected 'Зона' + 'Балл')")

    # find question col = where header contains 'Зона'
    q_col = None
    for idx, h in enumerate(header):
        if "зона" in (h or "").strip().lower():
            q_col = idx
            break
    if q_col is None:
        raise HTTPException(status_code=400, detail="Cannot detect 'Зона' column")

    # if there is a leading blank column before 'Зона' => it's section col
    section_col = q_col - 1 if q_col > 0 and not (_cell_str(header[q_col - 1]) or "").strip() else None

    current_section = ""
    out: list[dict] = []

    for excel_row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row or not any(_cell_str(c) for c in row):
            continue

        row_list = list(row)
        section_cell = _cell_str(row_list[section_col]) if (section_col is not None and section_col < len(row_list)) else ""
        q_cell = _cell_str(row_list[q_col]) if (q_col is not None and q_col < len(row_list)) else ""

        # section row: section filled, question empty
        if section_col is not None and section_cell and not q_cell:
            current_section = section_cell
            continue

        # 3-col variant: treat row with only first col filled as section
        if section_col is None and q_cell and all(not _cell_str(c) for c in row_list[q_col+1:]):
            current_section = q_cell
            continue

        if not q_cell:
            continue

        section = current_section or section_cell

        out.append({
            "section": section,
            "text": q_cell,
            "answer_type": "yesno_score",
            "options": {"yes_label": "Да", "no_label": "Нет", "yes_score": 1, "no_score": 0},
            "is_required": False,
            "allow_comment": True,
            "allow_photos": True,
            "order": 0,
        })

    return None, out


async def _audit_next_version(db: AsyncSession, organization_id: int, name: str, scope: str, location_type: str | None) -> int:
    if ChecklistTemplate is None:
        raise HTTPException(status_code=500, detail="ChecklistTemplate model is not available")

    q = select(func.max(ChecklistTemplate.version)).where(
        ChecklistTemplate.organization_id == int(organization_id),
        ChecklistTemplate.name == str(name),
        ChecklistTemplate.scope == str(scope),
        ChecklistTemplate.location_type == (str(location_type) if location_type else None),
    )
    max_v = (await db.execute(q)).scalar_one_or_none()
    return int(max_v or 0) + 1


@router.get("/organizations/{organization_id}/audit-templates-import/template")
async def download_audit_templates_import_template(
    organization_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.director, Role.super_admin)),
):
    await _ensure_import_access(db=db, user=user, organization_id=organization_id)

    try:
        from openpyxl import Workbook  # lazy import
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta.append(["key", "value"])
    ws_meta.append(["name", "HSK — чек-лист (пример)"])
    ws_meta.append(["scope", "organization"])   # organization | group | location
    ws_meta.append(["location_type", ""])       # optional, for scope=group
    ws_meta.append(["description", ""])

    ws_q = wb.create_sheet("questions")
    ws_q.append(["section", "question", "answer_type", "required", "allow_comment", "allow_photos", "order", "options_json"])
    ws_q.append(["Лобби", "Лобби чистое и опрятное", "yesno_score", "0", "1", "1", "1", ""])
    ws_q.append(["Лобби", "Стойка регистрации без мусора", "yesno_score", "0", "1", "1", "2", ""])

    buf = io.BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="audit_templates_import_template.xlsx"'},
    )


@router.post("/organizations/{organization_id}/audit-templates-import")
async def import_audit_templates_from_excel(
    organization_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.director, Role.super_admin)),
):
    await _ensure_import_access(db=db, user=user, organization_id=organization_id)

    if ChecklistTemplate is None or ChecklistQuestion is None:
        raise HTTPException(status_code=500, detail="Checklist models are not available (check imports in admin.py)")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        from openpyxl import load_workbook  # lazy import
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    try:
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid xlsx file")

    meta = _audit_parse_meta(wb)
    name = (meta.get("name") or "").strip() or (getattr(file, "filename", "") or "").strip() or (wb.active.title or "Checklist")
    scope = (meta.get("scope") or "organization").strip().lower()
    location_type = (meta.get("location_type") or "").strip() or None
    description = (meta.get("description") or "").strip() or None

    # Parse questions
    ws_q = _audit_find_questions_sheet(wb)
    if ws_q is not None:
        _, questions = _audit_parse_structured_questions(ws_q)
    else:
        # fallback: Word-like exported table
        _, questions = _audit_parse_hsk_like(wb.active)

    if not questions:
        raise HTTPException(status_code=400, detail="No questions found in xlsx")

    new_version = await _audit_next_version(db=db, organization_id=organization_id, name=name, scope=scope, location_type=location_type)

    # deactivate previous active templates for same (org+name+scope+location_type)
    await db.execute(
        update(ChecklistTemplate)
        .where(
            ChecklistTemplate.organization_id == int(organization_id),
            ChecklistTemplate.name == str(name),
            ChecklistTemplate.scope == str(scope),
            ChecklistTemplate.location_type == (str(location_type) if location_type else None),
            ChecklistTemplate.is_active == True,  # noqa: E712
        )
        .values(is_active=False)
    )

    tmpl = ChecklistTemplate(
        organization_id=int(organization_id),
        name=str(name)[:200],
        description=description,
        scope=str(scope),
        location_type=location_type,
        version=int(new_version),
        is_active=True,
        created_at=datetime.now(timezone.utc),
    )
    db.add(tmpl)
    await db.flush()  # get tmpl.id

    # normalize ordering
    # if order is not set -> sequential by appearance
    seq = 1
    for item in questions:
        order = int(item.get("order") or 0)
        if order <= 0:
            order = seq
        seq += 1

        q = ChecklistQuestion(
            template_id=int(tmpl.id),
            order=int(order),
            section=(item.get("section") or "")[:200],
            text=(item.get("text") or "")[:2000],
            answer_type=(item.get("answer_type") or "yesno_score")[:50],
            options=item.get("options"),
            is_required=bool(item.get("is_required") or False),
            allow_comment=bool(item.get("allow_comment") if item.get("allow_comment") is not None else True),
            allow_photos=bool(item.get("allow_photos") if item.get("allow_photos") is not None else True),
        )
        db.add(q)

    await db.commit()

    return {
        "ok": True,
        "organization_id": int(organization_id),
        "template_id": int(tmpl.id),
        "name": tmpl.name,
        "scope": tmpl.scope,
        "location_type": tmpl.location_type,
        "version": int(tmpl.version),
        "questions_count": int(len(questions)),
    }


@router.patch("/locations/{location_id}")
async def update_location(
    location_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles(Role.director)),
):
    """
    payload (any subset):
      { "name": "...", "type": "...", "code": "...", "is_active": true|false }
    """
    loc = (await db.execute(select(Location).where(Location.id == location_id))).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    if "name" in payload:
        name = (payload.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="name cannot be empty")
        loc.name = name

    if "type" in payload:
        loc.type = (payload.get("type") or "room").strip() or "room"

    if "code" in payload:
        loc.code = (payload.get("code") or "").strip()

    if "is_active" in payload:
        loc.is_active = bool(payload.get("is_active"))

    db.add(loc)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Location slug already exists")

    await db.refresh(loc)
    return {
        "id": loc.id,
        "organization_id": loc.organization_id,
        "type": loc.type,
        "code": loc.code,
        "name": loc.name,
        "slug": loc.slug,
        "is_active": loc.is_active,
    }



@router.post("/locations/bulk-rooms")
async def bulk_create_rooms(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.service_manager, Role.director)),
):
    """
    Массовое создание номеров (эффектно на демо).
    payload:
    {
      "organization_id": 1,
      "from": 101,
      "to": 120,
      "slug_prefix": "room-",
      "name_prefix": "Номер ",
      "type": "room"
    }
    """
    try:
        organization_id = int(payload["organization_id"])
        start = int(payload["from"])
        end = int(payload["to"])
    except Exception:
        raise HTTPException(status_code=400, detail="organization_id, from, to are required")

    if start > end or start <= 0:
        raise HTTPException(status_code=400, detail="Invalid range")

    allowed = await get_allowed_organization_ids(db=db, user=user)
    if organization_id not in allowed:
        raise HTTPException(status_code=403, detail="No access to this organization")

    slug_prefix = (payload.get("slug_prefix") or "room-").strip()
    name_prefix = (payload.get("name_prefix") or "Номер ").strip()
    loc_type = (payload.get("type") or "room").strip()

    created = 0
    skipped = 0

    for num in range(start, end + 1):
        slug = f"{slug_prefix}{num}"
        name = f"{name_prefix}{num}"
        loc = Location(
            organization_id=organization_id,
            type=loc_type,
            code=str(num),
            name=name,
            slug=slug,
            is_active=True,
        )
        db.add(loc)
        try:
            await db.flush()  # ловим уникальные конфликты в цикле
            created += 1
        except Exception:
            await db.rollback()
            skipped += 1
            continue

    await db.commit()
    return {"ok": True, "created": created, "skipped": skipped}


@router.patch("/locations/{location_id}/deactivate")
async def deactivate_location(
    location_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
    Role.admin,
    Role.ops_director,
    Role.manager,
    Role.service_manager,
    Role.director,
)),
):
    loc = (await db.execute(select(Location).where(Location.id == location_id))).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    allowed = await get_allowed_organization_ids(db=db, user=user)
    if loc.organization_id not in allowed:
        raise HTTPException(status_code=403, detail="No access to this location")

    loc.is_active = False
    db.add(loc)
    await db.commit()
    return {"ok": True, "id": loc.id, "is_active": loc.is_active}


@router.patch("/locations/{location_id}/activate")
async def activate_location(
    location_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
    )),
):
    loc = (await db.execute(select(Location).where(Location.id == location_id))).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    allowed = await get_allowed_organization_ids(db=db, user=user)
    if loc.organization_id not in allowed:
        raise HTTPException(status_code=403, detail="No access to this location")

    loc.is_active = True
    db.add(loc)
    await db.commit()
    return {"ok": True, "id": loc.id, "is_active": loc.is_active}


@router.get("/locations/{location_id}/submissions")
async def list_submissions(
    location_id: int,
    limit: int = 20,
    offset: int = 0,
    created_from: datetime | None = None,
    created_to: datetime | None = None,
    rating_min: int | None = None,
    rating_max: int | None = None,
    has_comment: bool | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.director,
        Role.auditor_global,
        Role.auditor,
        Role.manager,
        Role.service_manager,
        Role.employee,
    )),
):
    # RBAC по location
    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    # Базовые выражения по JSON answers
    rating_expr = cast(func.nullif(Submission.answers["rating_overall"].astext, ""), Integer)
    comment_expr = func.coalesce(Submission.answers["comment"].astext, "")

    where = [Submission.location_id == location_id]

    if created_from is not None:
        where.append(Submission.created_at >= created_from)
    if created_to is not None:
        where.append(Submission.created_at <= created_to)

    if rating_min is not None:
        where.append(rating_expr >= int(rating_min))
    if rating_max is not None:
        where.append(rating_expr <= int(rating_max))

    if has_comment is True:
        where.append(func.length(func.trim(comment_expr)) > 0)
    elif has_comment is False:
        where.append(func.length(func.trim(comment_expr)) == 0)

    # total
    total = (
        await db.execute(
            select(func.count(Submission.id)).where(*where)
        )
    ).scalar_one()

    # items
    rows = (
        await db.execute(
            select(Submission)
            .where(*where)
            .order_by(desc(Submission.created_at), desc(Submission.id))
            .offset(offset)
            .limit(min(max(limit, 1), 100))
        )
    ).scalars().all()

    items = []
    for s in rows:
        a = s.answers or {}
        m = s.meta or {}

        items.append(
            {
                "id": s.id,
                "location_id": s.location_id,
                "survey_version_id": s.survey_version_id,
                "created_at": s.created_at.isoformat(),
                "rating_overall": a.get("rating_overall"),
                "comment": a.get("comment") or "",
                "name": a.get("name") or "",
                "email": a.get("email") or "",

                # Patch 8.4.1: гостевой контекст (из meta)
                "room": m.get("room") or "",
                "guest_name": m.get("guest_name") or "",
                "reservation_code": m.get("reservation_code"),
                "stay_id": m.get("stay_id"),
            }
        )

    return {"total": int(total), "limit": int(limit), "offset": int(offset), "items": items}


@router.get("/submissions/{submission_id}")
async def get_submission(
    submission_id: int,
    dev: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            Role.admin,
            Role.ops_director,
            Role.director,
            Role.auditor_global,
            Role.auditor,
            Role.manager,
            Role.service_manager,
            Role.employee,
        )
    ),
):
    s = (
        await db.execute(select(Submission).where(Submission.id == submission_id))
    ).scalar_one_or_none()

    if not s:
        raise HTTPException(status_code=404, detail="Submission not found")

    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if s.location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this submission")

    a = s.answers or {}
    m = s.meta or {}

     # PATCH C: кто/когда обновлял "принятые меры"
    service_action_updated_by_name = None
    if getattr(s, "service_action_updated_by", None):
        u = (
            await db.execute(select(User).where(User.id == s.service_action_updated_by))
        ).scalar_one_or_none()
        if u:
            service_action_updated_by_name = u.full_name or u.email

    # Dev mode is allowed only for Role.admin.
    dev_enabled = bool(dev) and await user_has_any_role(db, user, {Role.admin.value})

    # In non-dev mode we do NOT return raw answers/meta.
    # Instead we provide a UI-friendly list of extra fields.
    sv_schema = (
        await db.execute(
            select(SurveyVersion.schema).where(SurveyVersion.id == s.survey_version_id)
        )
    ).scalar_one_or_none() or {}

    skip_keys = {
        "rating_overall",
        "comment",
        "name",
        "email",
        "phone",
        "tel",
        "telephone",
    }
    extra_fields = humanize_extra_answers(
        schema=sv_schema,
        answers=a,
        dev=dev_enabled,
        hide_unknown=not dev_enabled,
        skip_keys=skip_keys,
    )

    phone_raw = a.get("phone") or a.get("tel") or a.get("telephone")
    phone = ""
    if phone_raw is not None:
        phone = str(phone_raw)

    guest_context = None
    if any(
        k in m
        for k in (
            "stay_id",
            "guest_name",
            "room",
            "checkin_at",
            "checkout_at",
            "reservation_code",
        )
    ):
        guest_context = {
            "stay_id": m.get("stay_id"),
            "guest_name": m.get("guest_name"),
            "room": m.get("room"),
            "checkin_at": m.get("checkin_at"),
            "checkout_at": m.get("checkout_at"),
            "reservation_code": m.get("reservation_code"),
            "stay_source": m.get("stay_source"),
        }

    service_action_updated_by_name = None
    if getattr(s, "service_action_updated_by", None):
        u = (
            await db.execute(
                select(User).where(User.id == s.service_action_updated_by)
            )
        ).scalar_one_or_none()
        if u:
            service_action_updated_by_name = u.full_name or u.email

    return {
        "id": s.id,
        "location_id": s.location_id,
        "survey_version_id": s.survey_version_id,
        "created_at": s.created_at.isoformat(),
        "dev_mode": dev_enabled,
        "answers": a if dev_enabled else None,
        "meta": m if dev_enabled else None,
        "extra_fields": extra_fields,
        "phone": phone,

        # Patch 8.4.1 (не ломает старых клиентов)
        "guest_context": guest_context,
        "room": m.get("room") or "",
        "guest_name": m.get("guest_name") or "",
        "reservation_code": m.get("reservation_code"),
        "stay_id": m.get("stay_id"),

        # PATCH C: ответ / принятые меры сервис-менеджера
        "service_action_comment": (getattr(s, "service_action_comment", None) or ""),
        "service_action_updated_at": (
            s.service_action_updated_at.isoformat()
            if getattr(s, "service_action_updated_at", None)
            else None
        ),
        "service_action_updated_by": getattr(s, "service_action_updated_by", None),
        "service_action_updated_by_name": service_action_updated_by_name,

        "rating_overall": a.get("rating_overall"),
        "comment": a.get("comment") or "",
        "name": a.get("name") or "",
        "email": a.get("email") or "",
    }


@router.patch("/submissions/{submission_id}/action-comment")
async def update_submission_action_comment(
    submission_id: int,
    payload: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.service_manager)),
):
    """
    payload:
      { "service_action_comment": "..." }

    Доступ:
      - admin / ops_director / service_manager
      - auditor: read-only (не имеет роли на PATCH)
    """
    s = (
        await db.execute(select(Submission).where(Submission.id == submission_id))
    ).scalar_one_or_none()
    if not s:
        raise HTTPException(status_code=404, detail="Submission not found")

    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if s.location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this submission")

    if "service_action_comment" not in payload:
        raise HTTPException(status_code=400, detail="service_action_comment is required")

    comment = payload.get("service_action_comment")
    if comment is None:
        comment = ""
    if not isinstance(comment, str):
        raise HTTPException(status_code=400, detail="service_action_comment must be a string")

    comment = comment.strip()
    if len(comment) > 5000:
        raise HTTPException(status_code=400, detail="service_action_comment is too long (max 5000)")

    s.service_action_comment = comment
    s.service_action_updated_at = datetime.now(timezone.utc)
    s.service_action_updated_by = user.id

    await db.commit()
    await db.refresh(s)

    return {
        "ok": True,
        "id": s.id,
        "service_action_comment": s.service_action_comment or "",
        "service_action_updated_at": s.service_action_updated_at.isoformat() if s.service_action_updated_at else None,
        "service_action_updated_by": s.service_action_updated_by,
        "service_action_updated_by_name": (user.full_name or user.email),
    }


# =========================
# Stays (Patch 8.2.1)
# =========================

def _decode_csv_bytes(data: bytes) -> tuple[str, str]:
    """Decode bytes to text, максимально терпимо.

    Порядок предпочтений:
      - utf-8-sig / utf-8 (строгий)
      - cp1251
      - latin-1 (как последняя попытка)
    """
    for enc in ("utf-8-sig", "utf-8"):
        try:
            return data.decode(enc), enc
        except UnicodeDecodeError:
            pass

    # cp1251 почти всегда "декодирует", но это ожидаемо для RU выгрузок
    try:
        return data.decode("cp1251"), "cp1251"
    except Exception:
        return data.decode("latin-1", errors="replace"), "latin-1"


def _detect_delimiter(sample: str) -> str:
    sample = sample or ""
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample, delimiters=";\t,|")
        return dialect.delimiter
    except Exception:
        first = sample.splitlines()[0] if sample.splitlines() else sample
        cands = [";", ",", "\t", "|"]
        best = max(cands, key=lambda d: first.count(d))
        return best


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("\ufeff", "")
    s = re.sub(r"[\s\-\/]+", "_", s)
    s = re.sub(r"[^a-z0-9а-яё_]+", "", s)
    s = re.sub(r"_{2,}", "_", s).strip("_")
    return s


def _parse_date_any(v: str | None) -> date | None:
    if v is None:
        return None
    s = (str(v) or "").strip()
    if not s:
        return None

    s = s.split("T")[0].split(" ")[0]

    fmts = [
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue

    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def _build_stay_field_map(headers: list[str]) -> dict[str, str]:
    """Map our canonical fields -> csv header key.

    Returns keys: room, guest_name, checkin, checkout, reservation_code
    """
    norm_to_raw = {_norm_header(h): h for h in headers}
    keys = set(norm_to_raw.keys())

    def pick(cands: set[str]) -> str | None:
        for c in cands:
            if c in keys:
                return norm_to_raw[c]
        return None

    room = pick({"room", "room_no", "roomnumber", "номер", "номеркомнаты", "комната", "rm", "номер_комнаты"})
    checkin = pick({"checkin", "arrival", "arrival_date", "arrive", "datein", "дата_заезда", "заезд"})
    checkout = pick({"checkout", "departure", "departure_date", "depart", "dateout", "дата_выезда", "выезд"})
    reservation_code = pick({"reservation_code", "reservation", "res", "resno", "booking", "booking_id", "confirmation", "folio", "folio_no", "код_брони", "номер_брони", "бронь"})

    guest_name = pick({"guest_name", "guest", "name", "fullname", "fio", "фио", "гость", "клиент", "guestfullname", "guestname"})

    last = pick({"last_name", "lastname", "surname", "фамилия"})
    first = pick({"first_name", "firstname", "имя"})
    middle = pick({"middle_name", "middlename", "patronymic", "отчество"})

    return {
        "room": room or "",
        "guest_name": guest_name or "",
        "guest_last": last or "",
        "guest_first": first or "",
        "guest_middle": middle or "",
        "checkin": checkin or "",
        "checkout": checkout or "",
        "reservation_code": reservation_code or "",
    }


@router.get("/locations/{location_id}/stays")
async def list_stays(
    location_id: int,
    room: str | None = None,
    q: str | None = None,
    on: date | None = None,
    limit: int = 50,
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.director, Role.auditor_global, Role.auditor, Role.manager, Role.service_manager)),
):
    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    where = [Stay.location_id == location_id]
    if room:
        where.append(Stay.room == room.strip())
    if q:
        where.append(func.lower(Stay.guest_name).like(f"%{q.strip().lower()}%"))
    if on:
        where.append(Stay.checkin_at <= on)
        where.append(Stay.checkout_at > on)

    total = (await db.execute(select(func.count(Stay.id)).where(*where))).scalar_one()

    rows = (
        await db.execute(
            select(Stay)
            .where(*where)
            .order_by(Stay.checkin_at.desc(), Stay.room.asc(), Stay.id.desc())
            .offset(max(int(offset), 0))
            .limit(min(max(int(limit), 1), 200))
        )
    ).scalars().all()

    items = []
    for s in rows:
        items.append(
            {
                "id": s.id,
                "location_id": s.location_id,
                "room": s.room,
                "guest_name": s.guest_name,
                "checkin_at": s.checkin_at.isoformat(),
                "checkout_at": s.checkout_at.isoformat(),
                "reservation_code": s.reservation_code,
                "source": s.source,
                "created_at": s.created_at.isoformat() if s.created_at else None,
            }
        )

    return {"total": int(total), "limit": int(limit), "offset": int(offset), "items": items}


@router.post("/locations/{location_id}/stays/import")
async def import_stays_csv(
    location_id: int,
    file: UploadFile = File(...),
    source: str = Query("csv", max_length=40),
    max_rows: int = Query(20000, ge=1, le=50000),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.director, Role.manager, Role.service_manager)),
):
    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    text, encoding = _decode_csv_bytes(raw)
    sample = "\n".join(text.splitlines()[:20])
    delimiter = _detect_delimiter(sample)

    buf = io.StringIO(text)
    reader = csv.reader(buf, delimiter=delimiter)

    first_row = None
    for row in reader:
        if any((c or "").strip() for c in row):
            first_row = row
            break

    if first_row is None:
        raise HTTPException(status_code=400, detail="No data rows")

    first_norm = [_norm_header(x) for x in first_row]
    looks_like_header = any(
        k in first_norm for k in ("room", "номер", "комната", "фио", "guest_name", "arrival", "дата_заезда")
    )

    inserted = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []

    if looks_like_header:
        headers = [c.strip() for c in first_row]
        field_map = _build_stay_field_map(headers)
        # важно: header уже прочитан csv.reader → указатель буфера стоит после header
        reader_iter = csv.DictReader(buf, fieldnames=headers, delimiter=delimiter)
    else:
        field_map = {
            "room": "__col0__",
            "guest_name": "__col1__",
            "checkin": "__col2__",
            "checkout": "__col3__",
            "reservation_code": "__col4__",
            "guest_last": "",
            "guest_first": "",
            "guest_middle": "",
        }
        reader_iter = itertools.chain([first_row], reader)

    def get_val(d: dict, key: str) -> str:
        v = d.get(key)
        return (str(v) if v is not None else "").strip()

    def build_guest_name(d: dict) -> str:
        if field_map.get("guest_name"):
            raw_name = get_val(d, field_map["guest_name"])
            if raw_name:
                return raw_name
        parts = []
        if field_map.get("guest_last"):
            parts.append(get_val(d, field_map["guest_last"]))
        if field_map.get("guest_first"):
            parts.append(get_val(d, field_map["guest_first"]))
        if field_map.get("guest_middle"):
            parts.append(get_val(d, field_map["guest_middle"]))
        return " ".join([p for p in parts if p]).strip()

    row_idx = 1

    async def upsert_one(room_v: str, guest_v: str, ci: date, co: date, res_code: str | None):
        nonlocal inserted, updated
        code = (res_code or "").strip() or None
        existing = None
        if code:
            existing = (
                await db.execute(
                    select(Stay).where(Stay.location_id == location_id, Stay.reservation_code == code)
                )
            ).scalar_one_or_none()
        else:
            existing = (
                await db.execute(
                    select(Stay).where(
                        Stay.location_id == location_id,
                        Stay.room == room_v,
                        Stay.guest_name == guest_v,
                        Stay.checkin_at == ci,
                        Stay.checkout_at == co,
                        Stay.reservation_code.is_(None),
                    )
                )
            ).scalar_one_or_none()

        if existing:
            existing.room = room_v
            existing.guest_name = guest_v
            existing.checkin_at = ci
            existing.checkout_at = co
            existing.reservation_code = code
            existing.source = source
            db.add(existing)
            updated += 1
        else:
            db.add(
                Stay(
                    location_id=location_id,
                    room=room_v,
                    guest_name=guest_v,
                    checkin_at=ci,
                    checkout_at=co,
                    reservation_code=code,
                    source=source,
                )
            )
            inserted += 1

    if looks_like_header:
        for d in reader_iter:
            if d is None:
                continue
            row_idx += 1
            if inserted + updated + skipped >= max_rows:
                break

            try:
                room_v = get_val(d, field_map["room"]) if field_map.get("room") else ""
                guest_v = build_guest_name(d)
                ci_s = get_val(d, field_map["checkin"]) if field_map.get("checkin") else ""
                co_s = get_val(d, field_map["checkout"]) if field_map.get("checkout") else ""
                res_code = get_val(d, field_map["reservation_code"]) if field_map.get("reservation_code") else ""

                if not room_v or not guest_v:
                    skipped += 1
                    continue

                ci = _parse_date_any(ci_s)
                co = _parse_date_any(co_s)
                if ci is None or co is None:
                    raise ValueError(f"Bad dates: checkin='{ci_s}', checkout='{co_s}'")
                if co < ci:
                    raise ValueError("checkout < checkin")

                await upsert_one(room_v, guest_v, ci, co, res_code)
            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})
    else:
        for row in reader_iter:
            if not row or not any((c or "").strip() for c in row):
                continue
            row_idx += 1
            if inserted + updated + skipped >= max_rows:
                break
            try:
                room_v = (row[0] or "").strip() if len(row) > 0 else ""
                guest_v = (row[1] or "").strip() if len(row) > 1 else ""
                ci_s = (row[2] or "").strip() if len(row) > 2 else ""
                co_s = (row[3] or "").strip() if len(row) > 3 else ""
                res_code = (row[4] or "").strip() if len(row) > 4 else ""

                if not room_v or not guest_v:
                    skipped += 1
                    continue
                ci = _parse_date_any(ci_s)
                co = _parse_date_any(co_s)
                if ci is None or co is None:
                    raise ValueError(f"Bad dates: checkin='{ci_s}', checkout='{co_s}'")
                if co < ci:
                    raise ValueError("checkout < checkin")
                await upsert_one(room_v, guest_v, ci, co, res_code)
            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})

    await db.commit()

    return {
        "ok": True,
        "location_id": location_id,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:200],
        "encoding": encoding,
        "delimiter": delimiter,
        "has_header": bool(looks_like_header),
        "max_rows": max_rows,
    }


@router.get("/locations/{location_id}/stays/export.csv")
async def export_stays_csv(
    location_id: int,
    template: bool = Query(False),
    room: str | None = None,
    q: str | None = None,
    on: date | None = None,
    max_rows: int = Query(50000, ge=1, le=50000),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.admin, Role.ops_director, Role.director, Role.auditor_global, Role.auditor, Role.manager, Role.service_manager)),
):
    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    headers = ["room", "guest_name", "checkin", "checkout", "reservation_code"]
    output = io.StringIO()
    w = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)
    w.writerow(headers)

    if not template:
        where = [Stay.location_id == location_id]
        if room:
            where.append(Stay.room == room.strip())
        if q:
            where.append(func.lower(Stay.guest_name).like(f"%{q.strip().lower()}%"))
        if on:
            where.append(Stay.checkin_at <= on)
            where.append(Stay.checkout_at > on)

        rows = (
            await db.execute(
                select(Stay)
                .where(*where)
                .order_by(Stay.checkin_at.desc(), Stay.room.asc(), Stay.id.desc())
                .limit(max_rows)
            )
        ).scalars().all()

        for s in rows:
            w.writerow(
                [
                    s.room,
                    s.guest_name,
                    s.checkin_at.isoformat(),
                    s.checkout_at.isoformat(),
                    s.reservation_code or "",
                ]
            )

    # Excel-friendly BOM
    data = ("\ufeff" + output.getvalue()).encode("utf-8")
    filename = f"stays_location_{location_id}.csv" if not template else "stays_template.csv"

    return Response(
        content=data,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )



# =========================
# Surveys (read-only, Patch 6.1)
# =========================

@router.get("/locations/{location_id}/surveys")
async def list_surveys_for_location(
    location_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),

    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
        Role.auditor_global,
        Role.auditor,
    )),
):
    # RBAC: non-global -> only allowed locations
    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)
    if not is_global:
        allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
        if location_id not in allowed_loc_ids:
            raise HTTPException(status_code=403, detail="No access to this location")

    loc = (await db.execute(select(Location).where(Location.id == location_id))).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    surveys = (
        await db.execute(
            select(Survey)
            .where(Survey.location_id == location_id)
            .order_by(Survey.is_archived.asc(), Survey.id.asc())
        )
    ).scalars().all()

    if not surveys:
        return []

    survey_ids = [s.id for s in surveys]

    # active versions (can be 0/1 per survey; if data corrupted -> pick latest)
    active_versions = (
        await db.execute(
            select(SurveyVersion)
            .where(
                SurveyVersion.survey_id.in_(survey_ids),
                SurveyVersion.is_active == True,  # noqa: E712
            )
        )
    ).scalars().all()

    active_by_survey: dict[int, SurveyVersion] = {}
    for v in active_versions:
        cur = active_by_survey.get(v.survey_id)
        if not cur or (v.version, v.created_at, v.id) > (cur.version, cur.created_at, cur.id):
            active_by_survey[v.survey_id] = v

    # updated_at per survey = max(version.created_at)
    upd_rows = (
        await db.execute(
            select(SurveyVersion.survey_id, func.max(SurveyVersion.created_at))
            .where(SurveyVersion.survey_id.in_(survey_ids))
            .group_by(SurveyVersion.survey_id)
        )
    ).all()
    updated_by = {int(sid): dt for sid, dt in upd_rows}

    # versions_count per survey
    cnt_rows = (
        await db.execute(
            select(SurveyVersion.survey_id, func.count(SurveyVersion.id))
            .where(SurveyVersion.survey_id.in_(survey_ids))
            .group_by(SurveyVersion.survey_id)
        )
    ).all()
    count_by = {int(sid): int(cnt) for sid, cnt in cnt_rows}

    out = []
    for s in surveys:
        av = active_by_survey.get(s.id)
        upd = updated_by.get(s.id)
        out.append(
            {
                "survey_id": s.id,
                "location_id": s.location_id,
                "name": s.name,
                "is_archived": bool(getattr(s, "is_archived", False)),
                "active_version": av.version if av else None,
                "active_version_id": av.id if av else None,
                "versions_count": count_by.get(s.id, 0),
                "updated_at": upd.isoformat() if upd else None,
            }
        )
    return out


@router.get("/surveys/{survey_id}")
async def get_survey_with_versions(
    survey_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
        Role.auditor_global,
        Role.auditor,
    )),
):
    survey = (await db.execute(select(Survey).where(Survey.id == survey_id))).scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")

    # RBAC by location
    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)
    if not is_global:
        allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
        if survey.location_id not in allowed_loc_ids:
            raise HTTPException(status_code=403, detail="No access to this survey")

    versions = (
        await db.execute(
            select(SurveyVersion)
            .where(SurveyVersion.survey_id == survey_id)
            .order_by(desc(SurveyVersion.version), desc(SurveyVersion.created_at), desc(SurveyVersion.id))
        )
    ).scalars().all()

    active_version_id = None
    for v in versions:
        if v.is_active:
            active_version_id = v.id
            break

    return {
        "id": survey.id,
        "location_id": survey.location_id,
        "name": survey.name,
        "is_archived": bool(getattr(survey, "is_archived", False)),
        "created_at": survey.created_at.isoformat() if survey.created_at else None,
        "active_version_id": active_version_id,
        "versions": [
            {
                "id": v.id,
                "version": v.version,
                "is_active": bool(v.is_active),
                "created_at": v.created_at.isoformat() if v.created_at else None,
            }
            for v in versions
        ],
    }


# =========================
# Survey Versions (edit JSON) - Patch 6.2
# =========================

def _validate_schema_minimal(schema: dict):
    if not isinstance(schema, dict):
        raise HTTPException(status_code=422, detail="schema must be an object")

    # мягкая базовая проверка: не ломаем гибкость
    if "title" not in schema:
        raise HTTPException(status_code=422, detail="schema.title is required")

    slides = schema.get("slides")
    if not isinstance(slides, list):
        raise HTTPException(status_code=422, detail="schema.slides must be an array")

    meta = schema.get("meta")
    if meta is not None and not isinstance(meta, dict):
        raise HTTPException(status_code=422, detail="schema.meta must be an object")


@router.get("/survey-versions/{version_id}")
async def get_survey_version(
    version_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
        Role.auditor_global,
        Role.auditor,
    )),
):
    # Получаем версию + контекст:
    # - для location survey: Survey.location_id
    # - для group survey: GroupSurveyBinding.organization_id + group_key (через survey_id)
    row = (
        await db.execute(
            select(
                SurveyVersion,
                Survey.location_id,
                GroupSurveyBinding.organization_id,
                GroupSurveyBinding.group_key,
            )
            .join(Survey, Survey.id == SurveyVersion.survey_id)
            .outerjoin(GroupSurveyBinding, GroupSurveyBinding.survey_id == Survey.id)
            .where(SurveyVersion.id == version_id)
        )
    ).first()

    if not row:
        raise HTTPException(status_code=404, detail="SurveyVersion not found")

    sv: SurveyVersion = row[0]
    location_id_raw = row[1]
    org_id_raw = row[2]
    group_key_raw = row[3]

    roles = (await db.execute(select(UserRole).where(UserRole.user_id == user.id))).scalars().all()
    is_global = any(r.role in GLOBAL_ROLE_VALUES for r in roles)

    organization_id: int | None = None
    group_key: str | None = None
    location_id: int | None = None

    if location_id_raw is not None:
        location_id = int(location_id_raw)

        # RBAC: non-global -> allowed_locations
        if not is_global:
            allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
            if location_id not in [int(x) for x in allowed_loc_ids]:
                raise HTTPException(status_code=403, detail="No access to this survey version")

    else:
        # group survey
        if org_id_raw is None or group_key_raw is None:
            raise HTTPException(status_code=400, detail="SurveyVersion has no location_id and no group binding")

        organization_id = int(org_id_raw)
        group_key = str(group_key_raw)

        # RBAC: non-global -> must have at least one allowed location in this org+group
        if not is_global:
            allowed_loc_ids = await _require_group_access(db, user, organization_id, group_key)
            if not allowed_loc_ids:
                raise HTTPException(status_code=403, detail="No access to this group survey version")

    return {
        "id": sv.id,
        "survey_id": sv.survey_id,
        "location_id": location_id,  # null for group survey
        "organization_id": organization_id,  # null for location survey
        "group_key": group_key,  # null for location survey
        "version": sv.version,
        "is_active": bool(sv.is_active),
        "schema": sv.schema,
        "widget_config": sv.widget_config,
        "created_at": sv.created_at.isoformat() if sv.created_at else None,
    }


@router.patch("/survey-versions/{version_id}")
async def update_survey_version(
    version_id: int,
    payload: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
    )),
):
    row = (
        await db.execute(
            select(
                SurveyVersion,
                Survey.location_id,
                GroupSurveyBinding.organization_id,
                GroupSurveyBinding.group_key,
            )
            .join(Survey, Survey.id == SurveyVersion.survey_id)
            .outerjoin(GroupSurveyBinding, GroupSurveyBinding.survey_id == Survey.id)
            .where(SurveyVersion.id == version_id)
        )
    ).first()

    if not row:
        raise HTTPException(status_code=404, detail="SurveyVersion not found")

    sv: SurveyVersion = row[0]
    location_id_raw = row[1]
    org_id_raw = row[2]
    group_key_raw = row[3]

    roles = (await db.execute(select(UserRole).where(UserRole.user_id == user.id))).scalars().all()
    is_global = any(r.role in GLOBAL_ROLE_VALUES for r in roles)

    if location_id_raw is not None:
        location_id = int(location_id_raw)

        if not is_global:
            allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
            if location_id not in [int(x) for x in allowed_loc_ids]:
                raise HTTPException(status_code=403, detail="No access to this survey version")

    else:
        if org_id_raw is None or group_key_raw is None:
            raise HTTPException(status_code=400, detail="SurveyVersion has no location_id and no group binding")

        organization_id = int(org_id_raw)
        group_key = str(group_key_raw)

        if not is_global:
            allowed_loc_ids = await _require_group_access(db, user, organization_id, group_key)
            if not allowed_loc_ids:
                raise HTTPException(status_code=403, detail="No access to this group survey version")

    if not isinstance(payload, dict):
        raise HTTPException(status_code=422, detail="payload must be an object")

    has_any = False

    if "schema" in payload:
        has_any = True
        schema = payload.get("schema")
        if not isinstance(schema, dict):
            raise HTTPException(status_code=422, detail="schema must be an object")
        _validate_schema_minimal(schema)
        sv.schema = schema

    if "widget_config" in payload:
        has_any = True
        wc = payload.get("widget_config")
        if not isinstance(wc, dict):
            raise HTTPException(status_code=422, detail="widget_config must be an object")
        sv.widget_config = wc

    if not has_any:
        raise HTTPException(status_code=422, detail="Nothing to update (schema/widget_config)")

    await db.commit()
    await db.refresh(sv)

    return {"ok": True, "id": sv.id}


# =========================
# Surveys: Versions + Active (PATCH 6.3)
# =========================

@router.post("/surveys/{survey_id}/versions")
async def create_survey_version_copy(
    survey_id: int,
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
    )),
):
    """
    Создаёт новую версию как копию:
      - если payload.from_version_id передан -> копируем её
      - иначе копируем активную версию survey (если есть)
      - иначе копируем самую свежую версию survey
    payload:
      { "from_version_id": 123, "make_active": false }
    """
    from_version_id = payload.get("from_version_id")
    make_active = bool(payload.get("make_active", False))

    survey = (await db.execute(select(Survey).where(Survey.id == survey_id))).scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")

    if bool(getattr(survey, "is_archived", False)):
        raise HTTPException(status_code=400, detail="Survey is archived")
    
    if survey.location_id is None:
        raise HTTPException(status_code=400, detail="This is a group survey. Use group endpoints.")

    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if survey.location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    # source version
    source: SurveyVersion | None = None
    if from_version_id:
        source = (
            await db.execute(
                select(SurveyVersion).where(
                    SurveyVersion.id == int(from_version_id),
                    SurveyVersion.survey_id == survey_id,
                )
            )
        ).scalar_one_or_none()
        if not source:
            raise HTTPException(status_code=404, detail="Source version not found")
    else:
        source = (
            await db.execute(
                select(SurveyVersion)
                .where(SurveyVersion.survey_id == survey_id, SurveyVersion.is_active == True)  # noqa: E712
                .order_by(desc(SurveyVersion.version), desc(SurveyVersion.id))
                .limit(1)
            )
        ).scalar_one_or_none()

        if not source:
            source = (
                await db.execute(
                    select(SurveyVersion)
                    .where(SurveyVersion.survey_id == survey_id)
                    .order_by(desc(SurveyVersion.version), desc(SurveyVersion.id))
                    .limit(1)
                )
            ).scalar_one_or_none()

        if not source:
            raise HTTPException(status_code=400, detail="No versions to copy")

    max_ver = (
        await db.execute(
            select(func.max(SurveyVersion.version)).where(SurveyVersion.survey_id == survey_id)
        )
    ).scalar_one()
    next_version = int(max_ver or 0) + 1

    now = datetime.now(timezone.utc)

    new_ver = SurveyVersion(
        survey_id=survey_id,
        version=next_version,
        is_active=False,
        schema=copy.deepcopy(source.schema),
        widget_config=copy.deepcopy(source.widget_config),
        created_at=now,  # ключевой фикс против NOT NULL / отсутствующего default
    )

    try:
        db.add(new_ver)
        await db.flush()  # получаем new_ver.id

        if make_active:
            # деактивируем ВСЕ версии по всей локации
            survey_ids = (
                await db.execute(select(Survey.id).where(Survey.location_id == survey.location_id))
            ).scalars().all()

            if survey_ids:
                await db.execute(
                    update(SurveyVersion)
                    .where(SurveyVersion.survey_id.in_(list(survey_ids)))
                    .values(is_active=False)
                )

            new_ver.is_active = True
            db.add(new_ver)

        await db.commit()
        await db.refresh(new_ver)

    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Failed to create version (conflict or missing defaults)") from e
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail="Failed to create version") from e

    return {
        "id": new_ver.id,
        "survey_id": new_ver.survey_id,
        "location_id": survey.location_id,
        "version": new_ver.version,
        "is_active": bool(new_ver.is_active),
    }


@router.post("/survey-versions/{version_id}/set-active")
async def set_active_survey_version(
    version_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
    )),
):
    """
    Делает версию активной. ВАЖНО: активная версия должна быть одна на локацию,
    поэтому перед активацией мы выключаем is_active у всех версий всех опросов этой локации.
    """
    row = (
        await db.execute(
            select(SurveyVersion, Survey)
            .join(Survey, Survey.id == SurveyVersion.survey_id)
            .where(SurveyVersion.id == version_id)
            .limit(1)
        )
    ).first()

    if not row:
        raise HTTPException(status_code=404, detail="Survey version not found")

    ver, survey = row

    if bool(getattr(survey, "is_archived", False)):
        raise HTTPException(status_code=400, detail="Survey is archived")

    if survey.location_id is None:
        raise HTTPException(status_code=400, detail="This is a group survey. Use group endpoints.")

    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if survey.location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    # 1) деактивируем ВСЕ версии по всей локации
    survey_ids = (
        await db.execute(select(Survey.id).where(Survey.location_id == survey.location_id))
    ).scalars().all()

    if survey_ids:
        await db.execute(
            update(SurveyVersion)
            .where(SurveyVersion.survey_id.in_(list(survey_ids)))
            .values(is_active=False)
        )

    # 2) активируем нужную
    ver.is_active = True
    db.add(ver)

    await db.commit()
    return {"ok": True, "version_id": ver.id, "survey_id": survey.id, "location_id": survey.location_id}


# =========================
# Surveys: create + archive (PATCH 6.5)
# =========================

def _default_schema(location_name: str | None = None) -> dict:
    return {
        "meta": {"version": 1},
        "title": location_name or "Оцените ваш опыт",
        "slides": [
            {
                "id": "s1",
                "type": "rating",
                "field": "rating_overall",
                "scale": 10,
                "title": "Как вам у нас?",
                "required": True,
            },
            {
                "id": "s2",
                "type": "text",
                "field": "comment",
                "title": "Что понравилось/что улучшить?",
                "required": False,
                "maxLength": 800,
            },
            {
                "id": "s3",
                "type": "contact",
                "title": "Контакт (если хотите)",
                "fields": [
                    {"type": "text", "field": "name", "required": False},
                    {"type": "email", "field": "email", "required": False},
                ],
            },
        ],
    }


def _default_widget_config() -> dict:
    return {
        "brand": {"radius": 14, "primary": "#6d28d9"},
        "texts": {"submit": "Отправить", "thanks": "Спасибо за отзыв!"},
    }

# =========================
# Group Surveys (PATCH G-2)
# =========================


_GROUP_LABELS_RU: dict[str, str] = {
    "room": "Номера",
    "restaurant": "Рестораны",
    "conference_hall": "Конференц-залы",
    "banquet_hall": "Банкетные залы",
    "other": "Другое",
}


def _group_label_ru(group_key: str) -> str:
    return _GROUP_LABELS_RU.get(group_key, group_key)


async def _require_group_access(
    db: AsyncSession, user: User, organization_id: int, group_key: str
) -> list[int]:
    """RBAC for group: return allowed location_ids inside (organization_id + group_key)."""
    return await require_group_access(db=db, user=user, organization_id=organization_id, group_key=group_key)

    is_global = await user_has_any_role(db, user, GLOBAL_ROLE_VALUES)
    if is_global:
        return loc_ids_int

    allowed = await get_allowed_location_ids(db=db, user=user)
    allowed_set = set(int(x) for x in allowed)
    return [lid for lid in loc_ids_int if lid in allowed_set]


async def _location_active_for_preview(db: AsyncSession, location_id: int):
    row = (
        await db.execute(
            select(SurveyVersion, Survey)
            .join(Survey, Survey.id == SurveyVersion.survey_id)
            .where(
                Survey.location_id == location_id,
                SurveyVersion.is_active == True,  # noqa: E712
                Survey.is_archived == False,      # noqa: E712
            )
            .order_by(
                desc(SurveyVersion.version),
                desc(SurveyVersion.created_at),
                desc(SurveyVersion.id),
            )
            .limit(1)
        )
    ).first()

    if not row:
        return None

    ver, survey = row
    return {
        "survey_id": survey.id,
        "version_id": ver.id,
        "version": ver.version,
        "schema": ver.schema,
        "widget_config": ver.widget_config,
    }


@router.get("/organizations/{organization_id}/groups/{group_key}/survey")
async def get_group_survey(
    organization_id: int,
    group_key: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            Role.auditor,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.auditor_global,
        )
    ),
):
    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    allowed_loc_ids = await _require_group_access(db, user, organization_id, group_key)
    if not allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this group")

    binding = (
        await db.execute(
            select(GroupSurveyBinding).where(
                GroupSurveyBinding.organization_id == organization_id,
                GroupSurveyBinding.group_key == group_key,
            )
        )
    ).scalar_one_or_none()

    out = {
        "organization_id": organization_id,
        "organization_name": org.name,
        "group_key": group_key,
        "group_name": _group_label_ru(group_key),
        "binding": None,
        "survey": None,
        "active": None,
        "versions": [],
        "fallback": None,
    }

    # fallback preview: first accessible location active survey (legacy)
    first_loc_id = int(allowed_loc_ids[0])
    first_loc = (await db.execute(select(Location).where(Location.id == first_loc_id))).scalar_one_or_none()
    fallback_active = await _location_active_for_preview(db, first_loc_id)
    if first_loc and fallback_active:
        out["fallback"] = {
            "mode": "location",
            "location_id": first_loc.id,
            "location_name": first_loc.name,
            "active": fallback_active,
        }

    if not binding:
        return out

    survey = (await db.execute(select(Survey).where(Survey.id == binding.survey_id))).scalar_one_or_none()
    if not survey:
        return out

    versions = (
        await db.execute(
            select(SurveyVersion)
            .where(SurveyVersion.survey_id == survey.id)
            .order_by(desc(SurveyVersion.version), desc(SurveyVersion.created_at), desc(SurveyVersion.id))
        )
    ).scalars().all()

    active_ver = None
    for v in versions:
        if v.id == binding.active_version_id:
            active_ver = v
            break

    out["binding"] = {
        "id": binding.id,
        "survey_id": binding.survey_id,
        "active_version_id": binding.active_version_id,
    }
    out["survey"] = {
        "id": survey.id,
        "name": survey.name,
        "is_archived": bool(getattr(survey, "is_archived", False)),
        "created_at": survey.created_at.isoformat() if survey.created_at else None,
    }
    out["versions"] = [
        {
            "id": v.id,
            "version": v.version,
            "is_active": bool(v.is_active),
            "created_at": v.created_at.isoformat() if v.created_at else None,
        }
        for v in versions
    ]

    if active_ver:
        out["active"] = {
            "survey_id": survey.id,
            "version_id": active_ver.id,
            "version": active_ver.version,
            "schema": active_ver.schema,
            "widget_config": active_ver.widget_config,
        }

    return out


@router.post("/organizations/{organization_id}/groups/{group_key}/survey/bootstrap")
async def bootstrap_group_survey(
    organization_id: int,
    group_key: str,
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.auditor_global,
        )
    ),
):
    org = (await db.execute(select(Organization).where(Organization.id == organization_id))).scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")

    allowed_loc_ids = await _require_group_access(db, user, organization_id, group_key)
    if not allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this group")

    exists = (
        await db.execute(
            select(GroupSurveyBinding.id).where(
                GroupSurveyBinding.organization_id == organization_id,
                GroupSurveyBinding.group_key == group_key,
            )
        )
    ).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=409, detail="Group survey already exists")

    name = (payload.get("name") or "").strip() or f"Опрос: {_group_label_ru(group_key)}"

    source_location_id = payload.get("source_location_id")
    if source_location_id is not None:
        try:
            source_location_id = int(source_location_id)
        except Exception:
            raise HTTPException(status_code=422, detail="source_location_id must be int")
        if source_location_id not in [int(x) for x in allowed_loc_ids]:
            raise HTTPException(status_code=403, detail="No access to source_location_id")
    else:
        source_location_id = int(allowed_loc_ids[0])

    source_loc = (await db.execute(select(Location).where(Location.id == int(source_location_id)))).scalar_one_or_none()
    if not source_loc:
        raise HTTPException(status_code=404, detail="Source location not found")

    # Try copy from source active version; fallback to default
    source_row = (
        await db.execute(
            select(SurveyVersion, Survey)
            .join(Survey, Survey.id == SurveyVersion.survey_id)
            .where(
                Survey.location_id == int(source_location_id),
                SurveyVersion.is_active == True,  # noqa: E712
                Survey.is_archived == False,      # noqa: E712
            )
            .order_by(desc(SurveyVersion.version), desc(SurveyVersion.created_at), desc(SurveyVersion.id))
            .limit(1)
        )
    ).first()

    if source_row:
        src_ver, _src_survey = source_row
        schema = copy.deepcopy(src_ver.schema)
        widget_config = copy.deepcopy(src_ver.widget_config)
    else:
        schema = _default_schema(source_loc.name)
        widget_config = _default_widget_config()

    now = datetime.now(timezone.utc)

    survey = Survey(
        location_id=None,
        name=name,
        is_archived=False,
        created_at=now,
    )
    db.add(survey)
    await db.flush()

    ver = SurveyVersion(
        survey_id=survey.id,
        version=1,
        is_active=True,
        schema=schema,
        widget_config=widget_config,
        created_at=now,
    )
    db.add(ver)
    await db.flush()

    binding = GroupSurveyBinding(
        organization_id=organization_id,
        group_key=group_key,
        survey_id=survey.id,
        active_version_id=ver.id,
    )
    db.add(binding)

    await db.commit()
    return {"ok": True, "organization_id": organization_id, "group_key": group_key, "survey_id": survey.id, "active_version_id": ver.id}


@router.post("/organizations/{organization_id}/groups/{group_key}/survey/activate")
async def activate_group_survey_version(
    organization_id: int,
    group_key: str,
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.auditor_global,
        )
    ),
):
    allowed_loc_ids = await _require_group_access(db, user, organization_id, group_key)
    if not allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this group")

    binding = (
        await db.execute(
            select(GroupSurveyBinding).where(
                GroupSurveyBinding.organization_id == organization_id,
                GroupSurveyBinding.group_key == group_key,
            )
        )
    ).scalar_one_or_none()
    if not binding:
        raise HTTPException(status_code=404, detail="Group survey not found")

    try:
        version_id = int(payload.get("version_id"))
    except Exception:
        raise HTTPException(status_code=422, detail="version_id is required")

    ver = (await db.execute(select(SurveyVersion).where(SurveyVersion.id == version_id))).scalar_one_or_none()
    if not ver:
        raise HTTPException(status_code=404, detail="Survey version not found")
    if ver.survey_id != binding.survey_id:
        raise HTTPException(status_code=400, detail="version_id does not belong to this group survey")

    await db.execute(
        update(SurveyVersion).where(SurveyVersion.survey_id == binding.survey_id).values(is_active=False)
    )

    ver.is_active = True
    binding.active_version_id = ver.id
    db.add(ver)
    db.add(binding)

    await db.commit()
    return {"ok": True, "version_id": ver.id, "survey_id": binding.survey_id, "organization_id": organization_id, "group_key": group_key}


@router.post("/organizations/{organization_id}/groups/{group_key}/survey/versions")
async def create_group_survey_version_copy(
    organization_id: int,
    group_key: str,
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(
        require_roles(
            # new roles
            Role.admin,
            Role.ops_director,
            Role.service_manager,
            # legacy roles (compat)
            Role.manager,
            Role.director,
            Role.auditor_global,
        )
    ),
):
    """Create a new version for group survey as a copy.

    payload:
      { "copy_from_version_id": 123 (optional), "make_active": false }
    """

    allowed_loc_ids = await _require_group_access(db, user, organization_id, group_key)
    if not allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this group")

    binding = (
        await db.execute(
            select(GroupSurveyBinding).where(
                GroupSurveyBinding.organization_id == organization_id,
                GroupSurveyBinding.group_key == group_key,
            )
        )
    ).scalar_one_or_none()
    if not binding:
        raise HTTPException(status_code=404, detail="Group survey not found")

    make_active = bool(payload.get("make_active", False))
    copy_from_version_id = payload.get("copy_from_version_id")

    source: SurveyVersion | None = None
    if copy_from_version_id is not None:
        try:
            vid = int(copy_from_version_id)
        except Exception:
            raise HTTPException(status_code=422, detail="copy_from_version_id must be int")
        source = (
            await db.execute(
                select(SurveyVersion).where(
                    SurveyVersion.id == vid,
                    SurveyVersion.survey_id == binding.survey_id,
                )
            )
        ).scalar_one_or_none()
        if not source:
            raise HTTPException(status_code=404, detail="Source version not found")
    else:
        source = (await db.execute(select(SurveyVersion).where(SurveyVersion.id == binding.active_version_id))).scalar_one_or_none()
        if not source:
            source = (
                await db.execute(
                    select(SurveyVersion)
                    .where(SurveyVersion.survey_id == binding.survey_id)
                    .order_by(desc(SurveyVersion.version), desc(SurveyVersion.id))
                    .limit(1)
                )
            ).scalar_one_or_none()
        if not source:
            raise HTTPException(status_code=400, detail="No versions to copy")

    max_ver = (
        await db.execute(select(func.max(SurveyVersion.version)).where(SurveyVersion.survey_id == binding.survey_id))
    ).scalar_one()
    next_version = int(max_ver or 0) + 1

    now = datetime.now(timezone.utc)

    new_ver = SurveyVersion(
        survey_id=binding.survey_id,
        version=next_version,
        is_active=False,
        schema=copy.deepcopy(source.schema),
        widget_config=copy.deepcopy(source.widget_config),
        created_at=now,
    )
    db.add(new_ver)
    await db.flush()

    if make_active:
        await db.execute(
            update(SurveyVersion).where(SurveyVersion.survey_id == binding.survey_id).values(is_active=False)
        )
        new_ver.is_active = True
        binding.active_version_id = new_ver.id
        db.add(new_ver)
        db.add(binding)

    await db.commit()
    return {"id": new_ver.id, "survey_id": binding.survey_id, "version": new_ver.version, "is_active": bool(new_ver.is_active)}


@router.post("/locations/{location_id}/surveys")
async def create_survey_for_location(
    location_id: int,
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
    )),
):
    """
    Создаёт Survey для локации + первую версию v1.

    payload:
      {
        "name": "Опрос ресторана",
        "copy_from_location_active": true,   # default true
        "make_active": false                 # default false
      }
    """
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="name is required")

    copy_from_location_active = payload.get("copy_from_location_active", True)
    make_active = bool(payload.get("make_active", False))

    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    loc = (await db.execute(select(Location).where(Location.id == location_id))).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    # Определяем schema/widget_config для первой версии
    schema = None
    widget_config = None

    if copy_from_location_active:
        # берём активную версию по локации (только не из архивных опросов)
        row = (
            await db.execute(
                select(SurveyVersion, Survey)
                .join(Survey, Survey.id == SurveyVersion.survey_id)
                .where(
                    Survey.location_id == location_id,
                    SurveyVersion.is_active == True,  # noqa: E712
                    Survey.is_archived == False,      # noqa: E712
                )
                .order_by(desc(SurveyVersion.created_at), desc(SurveyVersion.id))
                .limit(1)
            )
        ).first()
        if row:
            sv, _s = row
            schema = copy.deepcopy(sv.schema)
            widget_config = copy.deepcopy(sv.widget_config)

    if schema is None:
        schema = _default_schema(loc.name)
    if widget_config is None:
        widget_config = _default_widget_config()

    now = datetime.now(timezone.utc)

    new_survey = Survey(
        location_id=location_id,
        name=name,
        is_archived=False,
        created_at=now,
    )

    try:
        db.add(new_survey)
        await db.flush()  # получить new_survey.id

        v1 = SurveyVersion(
            survey_id=new_survey.id,
            version=1,
            is_active=False,
            schema=schema,
            widget_config=widget_config,
            created_at=now,
        )
        db.add(v1)
        await db.flush()

        if make_active:
            # деактивируем ВСЕ версии по всей локации
            survey_ids = (
                await db.execute(select(Survey.id).where(Survey.location_id == location_id))
            ).scalars().all()

            if survey_ids:
                await db.execute(
                    update(SurveyVersion)
                    .where(SurveyVersion.survey_id.in_(list(survey_ids)))
                    .values(is_active=False)
                )

            v1.is_active = True
            db.add(v1)

        await db.commit()
        await db.refresh(new_survey)
        await db.refresh(v1)

    except IntegrityError as e:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Failed to create survey") from e

    return {
        "survey_id": new_survey.id,
        "location_id": location_id,
        "name": new_survey.name,
        "is_archived": bool(new_survey.is_archived),
        "version_id": v1.id,
        "version": v1.version,
        "is_active": bool(v1.is_active),
    }


@router.patch("/surveys/{survey_id}")
async def update_survey(
    survey_id: int,
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(
        Role.admin,
        Role.ops_director,
        Role.manager,
        Role.service_manager,
        Role.director,
    )),
):
    """
    payload:
      { "name": "New name", "is_archived": true/false }
    """
    survey = (await db.execute(select(Survey).where(Survey.id == survey_id))).scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")

    allowed_loc_ids = await get_allowed_location_ids(db=db, user=user)
    if survey.location_id not in allowed_loc_ids:
        raise HTTPException(status_code=403, detail="No access to this location")

    changed = False

    if "name" in payload:
        name = (payload.get("name") or "").strip()
        if not name:
            raise HTTPException(status_code=422, detail="name cannot be empty")
        survey.name = name
        changed = True

    if "is_archived" in payload:
        is_archived = bool(payload.get("is_archived"))
        survey.is_archived = is_archived
        changed = True

        if is_archived:
            # при архивировании выключаем активность у версий этого survey,
            # чтобы public/resolve гарантированно его не выбирал
            await db.execute(
                update(SurveyVersion)
                .where(SurveyVersion.survey_id == survey.id)
                .values(is_active=False)
            )

    if not changed:
        raise HTTPException(status_code=422, detail="Nothing to update")

    await db.commit()
    return {"ok": True, "id": survey.id, "is_archived": bool(survey.is_archived), "name": survey.name}


# =========================
# Users assignments (director only) — Patch 7.2
# =========================

@router.post("/users/{user_id}/service-manager/{location_id}")
async def assign_service_manager_to_location(
    user_id: int,
    location_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles(Role.director)),
):
    """
    Назначить пользователю роль service_manager на конкретную локацию.
    Director-only.

    Идемпотентно:
      - если роль уже есть -> ok + status=already_assigned
    Дополнительно:
      - гарантируем доступ к организации через user_organizations (иначе у scoped юзера /admin/organizations будет пустой)
    """
    target = (await db.execute(select(User).where(User.id == int(user_id)))).scalar_one_or_none()
    if not target or not target.is_active:
        raise HTTPException(status_code=404, detail="User not found")

    loc = (
        await db.execute(
            select(Location)
            .join(Organization, Organization.id == Location.organization_id)
            .where(
                Location.id == int(location_id),
                Location.is_active == True,      # noqa: E712
                Organization.is_active == True,  # noqa: E712
            )
        )
    ).scalar_one_or_none()
    if not loc:
        raise HTTPException(status_code=404, detail="Location not found")

    # 1) ensure org access link (user_organizations)
    link = (
        await db.execute(
            select(UserOrganization).where(
                UserOrganization.user_id == target.id,
                UserOrganization.organization_id == loc.organization_id,
            )
        )
    ).scalar_one_or_none()

    if link:
        if not link.is_active:
            link.is_active = True
            db.add(link)
    else:
        db.add(
            UserOrganization(
                user_id=target.id,
                organization_id=loc.organization_id,
                is_active=True,
            )
        )

    # 2) ensure role row (user_roles)
    exists = (
        await db.execute(
            select(UserRole).where(
                UserRole.user_id == target.id,
                UserRole.role == Role.service_manager.value,
                UserRole.location_id == loc.id,
            )
        )
    ).scalar_one_or_none()

    if exists:
        await db.commit()
        return {
            "ok": True,
            "status": "already_assigned",
            "user_id": target.id,
            "location_id": loc.id,
        }

    db.add(
        UserRole(
            user_id=target.id,
            organization_id=loc.organization_id,
            location_id=loc.id,
            role=Role.service_manager.value,
        )
    )

    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        # на всякий случай считаем идемпотентным (например, гонка)
        return {
            "ok": True,
            "status": "already_assigned",
            "user_id": target.id,
            "location_id": loc.id,
        }

    return {
        "ok": True,
        "status": "assigned",
        "user_id": target.id,
        "location_id": loc.id,
    }


@router.delete("/users/{user_id}/service-manager/{location_id}")
async def remove_service_manager_from_location(
    user_id: int,
    location_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles(Role.director)),
):
    """
    Снять у пользователя роль service_manager с конкретной локации.
    Director-only.

    Идемпотентно:
      - если роли нет -> ok + status=not_assigned
    """
    target = (await db.execute(select(User).where(User.id == int(user_id)))).scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    role_row = (
        await db.execute(
            select(UserRole).where(
                UserRole.user_id == target.id,
                UserRole.role == Role.service_manager.value,
                UserRole.location_id == int(location_id),
            )
        )
    ).scalar_one_or_none()

    if not role_row:
        return {"ok": True, "status": "not_assigned", "user_id": target.id, "location_id": int(location_id)}

    await db.delete(role_row)
    await db.commit()
    return {"ok": True, "status": "removed", "user_id": target.id, "location_id": int(location_id)}
