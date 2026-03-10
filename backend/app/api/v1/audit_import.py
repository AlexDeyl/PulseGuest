from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile, Form
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.deps import get_db, get_current_user
from app.models.role import Role
from app.models.user import User
from app.services.rbac import require_org_access, require_roles

# ---- IMPORTANT: adjust import path if your models are in another file ----
try:
    from app.models.audit_checklist import ChecklistTemplate, ChecklistQuestion  # type: ignore
except Exception:
    ChecklistTemplate = None  # type: ignore
    ChecklistQuestion = None  # type: ignore

router = APIRouter()


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _audit_bool(v: Any, default: bool = False) -> bool:
    s = _cell_str(v).lower()
    if s in {"1", "true", "yes", "y", "on", "да"}:
        return True
    if s in {"0", "false", "no", "n", "off", "нет"}:
        return False
    return default


def _audit_int(v: Any, default: int = 0) -> int:
    s = _cell_str(v)
    if not s:
        return default
    try:
        return int(float(s))
    except Exception:
        return default


def _normalize_template_scope(v: Any) -> str:
    s = _cell_str(v).lower()
    if s in {"", "organization", "org", "организация"}:
        return "organization"
    if s in {"group", "location", "group_locations", "группа", "группа_локаций", "локации"}:
        # legacy: location -> теперь считаем group
        return "group"
    raise HTTPException(status_code=400, detail="Invalid meta.scope. Allowed: organization | group")


def _set_if_has(obj: Any, **kwargs: Any) -> None:
    for k, v in kwargs.items():
        if hasattr(obj, k):
            setattr(obj, k, v)


def _parse_meta(wb) -> dict[str, str]:
    meta: dict[str, str] = {}
    for name in ("meta", "Meta", "META", "Мета", "мета"):
        if name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 2:
                    continue
                k = _cell_str(row[0]).lower()
                v = _cell_str(row[1])
                if k:
                    meta[k] = v
            break
    return meta


def _find_questions_sheet(wb):
    for name in ("questions", "Questions", "Вопросы", "вопросы"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def _parse_questions(ws) -> list[dict[str, Any]]:
    """
    Expected columns (any aliases):
      section/раздел/зона | question/вопрос | answer_type/тип | options_json/опции/варианты
      required/обязателен | allow_comment/комментарий | allow_photos/фото | order/порядок/№
    """
    header_row_idx = None
    header_values = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(_cell_str(c) for c in row):
            header_row_idx = i
            header_values = [(_cell_str(c) or "") for c in row]
            break

    if header_row_idx is None or header_values is None:
        raise HTTPException(status_code=400, detail="No header row found in xlsx")

    aliases = {
        "section": "section",
        "раздел": "section",
        "зона": "section",

        "question": "question",
        "вопрос": "question",
        "пункт": "question",

        "answer_type": "answer_type",
        "type": "answer_type",
        "тип": "answer_type",

        "options_json": "options_json",
        "options": "options_json",
        "опции": "options_json",
        "варианты": "options_json",

        "required": "required",
        "обязателен": "required",
        "обязательно": "required",

        "allow_comment": "allow_comment",
        "комментарий": "allow_comment",

        "allow_photos": "allow_photos",
        "фото": "allow_photos",

        "order": "order",
        "порядок": "order",
        "№": "order",
        "#": "order",
    }

    norm = []
    for h in header_values:
        key = (h or "").strip().lower().replace("-", "_").replace(" ", "_")
        norm.append(aliases.get(key, key))

    def col(name: str) -> int | None:
        try:
            return norm.index(name)
        except ValueError:
            return None

    i_question = col("question")
    if i_question is None:
        raise HTTPException(status_code=400, detail="Missing column: question/вопрос")

    i_section = col("section")
    i_answer_type = col("answer_type")
    i_options = col("options_json")
    i_required = col("required")
    i_allow_comment = col("allow_comment")
    i_allow_photos = col("allow_photos")
    i_order = col("order")

    out: list[dict[str, Any]] = []
    for excel_row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not row or not any(_cell_str(c) for c in row):
            continue

        q_text = _cell_str(row[i_question] if i_question < len(row) else "")
        if not q_text:
            continue

        section = _cell_str(row[i_section] if (i_section is not None and i_section < len(row)) else "")
        answer_type = _cell_str(row[i_answer_type] if (i_answer_type is not None and i_answer_type < len(row)) else "") or "yesno_score"
        order = _audit_int(row[i_order] if (i_order is not None and i_order < len(row)) else "", default=0)

        is_required = _audit_bool(row[i_required] if (i_required is not None and i_required < len(row)) else "", default=False)
        allow_comment = _audit_bool(row[i_allow_comment] if (i_allow_comment is not None and i_allow_comment < len(row)) else "", default=True)
        allow_photos = _audit_bool(row[i_allow_photos] if (i_allow_photos is not None and i_allow_photos < len(row)) else "", default=True)

        options: Any = None
        if i_options is not None and i_options < len(row):
            raw_opt = _cell_str(row[i_options])
            if raw_opt:
                try:
                    options = json.loads(raw_opt)
                except Exception:
                    raise HTTPException(status_code=400, detail=f"Invalid options_json at row {excel_row_idx}")

        out.append(
            {
                "section": section,
                "text": q_text,
                "answer_type": answer_type,
                "options": options,
                "is_required": is_required,
                "allow_comment": allow_comment,
                "allow_photos": allow_photos,
                "order": order,
            }
        )

    if not out:
        raise HTTPException(status_code=400, detail="No questions found in xlsx")
    return out


async def _next_version(db: AsyncSession, organization_id: int, name: str, scope: str, location_type: str | None) -> int:
    q = select(func.max(ChecklistTemplate.version)).where(  # type: ignore
        ChecklistTemplate.organization_id == int(organization_id),  # type: ignore
        ChecklistTemplate.name == str(name),  # type: ignore
        ChecklistTemplate.scope == str(scope),  # type: ignore
        ChecklistTemplate.location_type == (str(location_type) if location_type else None),  # type: ignore
    )
    max_v = (await db.execute(q)).scalar_one_or_none()
    return int(max_v or 0) + 1


@router.get("/templates-import/template")
async def download_audit_import_template(
    organization_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.auditor, Role.auditor_global)),
):
    # RBAC: auditor only inside his orgs
    await require_org_access(db=db, user=user, organization_id=int(organization_id))

    from openpyxl import Workbook

    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta.append(["key", "value"])
    ws_meta.append(["name", "HSK — чек-лист (пример)"])
    ws_meta.append(["scope", "organization"])  # organization | group | location
    ws_meta.append(["location_type", ""])      # optional if scope=group/location
    ws_meta.append(["description", ""])

    ws_q = wb.create_sheet("questions")
    ws_q.append(["section", "question", "answer_type", "required", "allow_comment", "allow_photos", "order", "options_json"])
    ws_q.append(["Лобби", "Пол чистый?", "yesno_score", "0", "1", "1", "1", ""])
    ws_q.append(["Лобби", "Есть мусор/пятна?", "yesno_score", "0", "1", "1", "2", ""])
    ws_q.append(["Итог", "Общий комментарий (как текст)", "text", "0", "1", "1", "3", ""])

    buf = io.BytesIO()
    wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="audit_import_template.xlsx"'},
    )


@router.post("/templates-import")
async def import_audit_templates_from_excel(
    organization_id: int,
    file: UploadFile = File(...),
    group_value: str | None = Form(default=None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    _=Depends(require_roles(Role.auditor, Role.auditor_global)),
):
    if ChecklistTemplate is None or ChecklistQuestion is None:
        raise HTTPException(status_code=500, detail="Checklist models not found (check import in audit_import.py)")

    # RBAC: auditor only inside his orgs
    await require_org_access(db=db, user=user, organization_id=int(organization_id))

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    from openpyxl import load_workbook

    try:
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid xlsx file")

    meta = _parse_meta(wb)
    name = (meta.get("name") or "").strip() or (file.filename or "Checklist").strip()
    scope = _normalize_template_scope(meta.get("scope") or "organization")
    location_type = (meta.get("location_type") or "").strip() or None

    # Явный override с UI страницы импорта:
    # - group_value пустой -> organization
    # - group_value заполнен -> group + location_type=group_value
    gv = (group_value or "").strip()
    if gv:
        scope = "group"
        location_type = gv
    else:
        scope = "organization"
        location_type = None
    description = (meta.get("description") or "").strip() or None

    ws = _find_questions_sheet(wb)
    questions = _parse_questions(ws)

    new_version = await _next_version(db=db, organization_id=int(organization_id), name=name, scope=scope, location_type=location_type)

    # deactivate previous active templates for same key
    await db.execute(
        update(ChecklistTemplate)  # type: ignore
        .where(
            ChecklistTemplate.organization_id == int(organization_id),  # type: ignore
            ChecklistTemplate.name == str(name),  # type: ignore
            ChecklistTemplate.scope == str(scope),  # type: ignore
            ChecklistTemplate.location_type == (str(location_type) if location_type else None),  # type: ignore
            ChecklistTemplate.is_active == True,  # noqa: E712
        )
        .values(is_active=False)
    )

    tmpl = ChecklistTemplate()  # type: ignore
    _set_if_has(
        tmpl,
        organization_id=int(organization_id),
        name=str(name)[:200],
        description=description,
        scope=str(scope),
        location_type=location_type,
        version=int(new_version),
        is_active=True,
        created_at=datetime.now(timezone.utc),
    )
    # optional: if model has created_by_user_id
    _set_if_has(tmpl, created_by_user_id=int(user.id))

    db.add(tmpl)
    await db.flush()  # tmpl.id

    seq = 1
    for item in questions:
        q = ChecklistQuestion()  # type: ignore
        order = int(item.get("order") or 0) or seq
        seq += 1

        _set_if_has(
            q,
            template_id=int(getattr(tmpl, "id")),
            order=int(order),
            section=str(item.get("section") or "")[:200],
            text=str(item.get("text") or "")[:2000],
            answer_type=str(item.get("answer_type") or "yesno_score")[:50],
            is_required=bool(item.get("is_required") or False),
            allow_comment=bool(item.get("allow_comment") if item.get("allow_comment") is not None else True),
            allow_photos=bool(item.get("allow_photos") if item.get("allow_photos") is not None else True),
        )
        # options may be JSON column named options / options_json
        if item.get("options") is not None:
            if hasattr(q, "options"):
                setattr(q, "options", item["options"])
            elif hasattr(q, "options_json"):
                setattr(q, "options_json", item["options"])

        db.add(q)

    await db.commit()

    return {
        "ok": True,
        "organization_id": int(organization_id),
        "template_id": int(getattr(tmpl, "id")),
        "name": getattr(tmpl, "name", name),
        "scope": getattr(tmpl, "scope", scope),
        "location_type": getattr(tmpl, "location_type", location_type),
        "version": int(getattr(tmpl, "version", new_version)),
        "questions_count": int(len(questions)),
    }
